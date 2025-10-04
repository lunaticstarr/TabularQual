from __future__ import annotations

from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path
import shutil

from .types import InMemoryModel, Person


def write_spreadsheet(model: InMemoryModel, output_path: str, template_path: str = None):
    """Write InMemoryModel to spreadsheet format"""
    
    # Use template if provided, otherwise create new workbook
    if template_path and Path(template_path).exists():
        # Copy template to output
        shutil.copy(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        
        # Remove data sheets if they exist, keep README and Appendix
        for sheet_name in ['Model', 'Species', 'Transitions', 'Interactions']:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    
    # Determine insertion position (after README if it exists)
    insert_pos = 1 if 'README' in wb.sheetnames else 0
    
    # Create sheets in order: README, Model, Species, Interactions, Transitions, Appendix
    _write_model_sheet(wb, model, insert_pos)
    _write_species_sheet(wb, model, insert_pos + 1)
    _write_interactions_sheet(wb, model, insert_pos + 2)
    _write_transitions_sheet(wb, model, insert_pos + 3)
    
    # Save
    wb.save(output_path)


def _write_model_sheet(wb: openpyxl.Workbook, model: InMemoryModel, position: int = 0):
    """Write Model sheet"""
    ws = wb.create_sheet("Model", position)
    
    # No headers for Model sheet
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Data rows (start from row 1)
    row = 1
    
    # Model_source
    if model.model.source_urls:
        ws.cell(row=row, column=1, value="Model_source")
        ws.cell(row=row, column=2, value=", ".join(_url_to_compact_id(url) for url in model.model.source_urls))
        row += 1
    
    # Model_ID (required)
    cell = ws.cell(row=row, column=1, value="Model_ID")
    cell.fill = required_fill
    ws.cell(row=row, column=2, value=model.model.model_id)
    row += 1
    
    # Name
    if model.model.name:
        ws.cell(row=row, column=1, value="Name")
        ws.cell(row=row, column=2, value=model.model.name)
        row += 1
    
    # Publication
    if model.model.described_by:
        ws.cell(row=row, column=1, value="Publication")
        ws.cell(row=row, column=2, value=", ".join(_url_to_compact_id(url) for url in model.model.described_by))
        row += 1
    
    # Origin_publication & Origin_model (from derived_from)
    origin_pubs = []
    origin_models = []
    if model.model.derived_from:
        for url in model.model.derived_from:
            compact_id = _url_to_compact_id(url)
            if 'biomodels' in compact_id.lower():
                origin_models.append(compact_id)
            else:
                origin_pubs.append(compact_id)
    
    if origin_pubs:
        ws.cell(row=row, column=1, value="Origin_publication")
        ws.cell(row=row, column=2, value=", ".join(origin_pubs))
        row += 1
    
    if origin_models:
        ws.cell(row=row, column=1, value="Origin_model")
        ws.cell(row=row, column=2, value=", ".join(origin_models))
        row += 1
    
    # Taxon
    if model.model.taxons:
        ws.cell(row=row, column=1, value="Taxon")
        ws.cell(row=row, column=2, value=", ".join(_url_to_compact_id(taxon) for taxon in model.model.taxons))
        row += 1
    
    # Biological_process
    if model.model.biological_processes:
        ws.cell(row=row, column=1, value="Biological_process")
        ws.cell(row=row, column=2, value=", ".join(_url_to_compact_id(process) for process in model.model.biological_processes))
        row += 1
    
    # Created
    if model.model.created_iso:
        ws.cell(row=row, column=1, value="Created")
        ws.cell(row=row, column=2, value=model.model.created_iso)
        row += 1
    
    # Modified - use current timestamp
    from datetime import datetime, timezone
    ws.cell(row=row, column=1, value="Modified")
    ws.cell(row=row, column=2, value=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"))
    row += 1
    
    # Creators
    for idx, creator in enumerate(model.model.creators, 1):
        ws.cell(row=row, column=1, value=f"Creator{idx}")
        ws.cell(row=row, column=2, value=_person_to_string(creator))
        row += 1
    
    # Contributors
    for idx, contributor in enumerate(model.model.contributors, 1):
        ws.cell(row=row, column=1, value=f"Contributor{idx}")
        ws.cell(row=row, column=2, value=_person_to_string(contributor))
        row += 1
    
    # Version
    if model.model.versions:
        for version in model.model.versions:
            ws.cell(row=row, column=1, value="Version")
            ws.cell(row=row, column=2, value=version)
            row += 1
    
    # Notes - combine all into single cell as Notes1
    if model.model.notes:
        ws.cell(row=row, column=1, value="Notes1")
        # Join all notes with double newline
        ws.cell(row=row, column=2, value="\n\n".join(model.model.notes))
        row += 1
    
    # Comments - add TabularQual version
    ws.cell(row=row, column=1, value="Comments")
    ws.cell(row=row, column=2, value="Created by TabularQual version 0.1.0")
    row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80


def _write_species_sheet(wb: openpyxl.Workbook, model: InMemoryModel, position: int = 0):
    """Write Species sheet"""
    ws = wb.create_sheet("Species", position)
    
    # Header style
    header_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    header_font = Font(bold=True)
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Group annotations by qualifier and determine max unique qualifiers
    max_qualifiers = 0
    max_notes = 0
    for species in model.species.values():
        # Group by qualifier
        grouped = {}
        for qualifier, identifier in species.annotations:
            if qualifier not in grouped:
                grouped[qualifier] = []
            grouped[qualifier].append(identifier)
        max_qualifiers = max(max_qualifiers, len(grouped))
        max_notes = max(max_notes, len(species.notes))
    
    # Build headers
    headers = ["Species_ID", "Name"]
    
    # Add Relation/Identifier pairs
    for i in range(max_qualifiers):
        headers.extend([f"Relation{i + 1}", f"Identifier{i + 1}"])
    
    headers.extend(["Compartment", "Type", "Constant", "InitialLevel", "MaxLevel"])
    
    # Add Notes
    for i in range(max_notes):
        headers.append(f"Notes{i + 1}")
    
    headers.append("Comments")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        if header == "Species_ID":
            cell.fill = required_fill
    
    # Write species data
    for row_idx, species in enumerate(sorted(model.species.values(), key=lambda s: s.species_id), 2):
        col = 1
        
        # Species_ID
        ws.cell(row=row_idx, column=col, value=species.species_id)
        col += 1
        
        # Name
        ws.cell(row=row_idx, column=col, value=species.name)
        col += 1
        
        # Annotations - group by qualifier and combine identifiers
        grouped_annos = {}
        for qualifier, identifier in species.annotations:
            if qualifier not in grouped_annos:
                grouped_annos[qualifier] = []
            grouped_annos[qualifier].append(_url_to_compact_id(identifier))
        
        for i in range(max_qualifiers):
            if i < len(grouped_annos):
                qualifier = list(grouped_annos.keys())[i]
                identifiers = grouped_annos[qualifier]
                ws.cell(row=row_idx, column=col, value=qualifier)
                ws.cell(row=row_idx, column=col + 1, value=", ".join(identifiers))
            col += 2
        
        # Compartment
        ws.cell(row=row_idx, column=col, value=species.compartment)
        col += 1
        
        # Type (not in our data structure, skip)
        col += 1
        
        # Constant
        if species.constant is not None:
            ws.cell(row=row_idx, column=col, value=str(species.constant))
        col += 1
        
        # InitialLevel
        if species.initial_level is not None:
            ws.cell(row=row_idx, column=col, value=species.initial_level)
        col += 1
        
        # MaxLevel
        if species.max_level is not None:
            ws.cell(row=row_idx, column=col, value=species.max_level)
        col += 1
        
        # Notes - write each note to separate column
        for i in range(max_notes):
            if i < len(species.notes):
                ws.cell(row=row_idx, column=col, value=species.notes[i])
            col += 1
        
        # Comments (skip)
        col += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30


def _write_transitions_sheet(wb: openpyxl.Workbook, model: InMemoryModel, position: int = 0):
    """Write Transitions sheet"""
    ws = wb.create_sheet("Transitions", position)
    
    # Header style
    header_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    header_font = Font(bold=True)
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Group annotations by qualifier and determine max unique qualifiers
    max_qualifiers = 0
    max_notes = 0
    for transition in model.transitions:
        # Group by qualifier
        grouped = {}
        for qualifier, identifier in transition.annotations:
            if qualifier not in grouped:
                grouped[qualifier] = []
            grouped[qualifier].append(identifier)
        max_qualifiers = max(max_qualifiers, len(grouped))
        max_notes = max(max_notes, len(transition.notes))
    
    # Build headers
    headers = ["Transition_ID", "Name", "Target", "Level", "Rule"]
    
    # Add Relation/Identifier pairs
    for i in range(max_qualifiers):
        headers.extend([f"Relation{i + 1}", f"Identifier{i + 1}"])
    
    # Add Notes
    for i in range(max_notes):
        headers.append(f"Notes{i + 1}")
    
    headers.append("Comments")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        if header in ["Target", "Rule"]:
            cell.fill = required_fill
    
    # Write transition data
    for row_idx, transition in enumerate(model.transitions, 2):
        col = 1
        
        # Transitions_ID
        ws.cell(row=row_idx, column=col, value=transition.transition_id)
        col += 1
        
        # Name
        ws.cell(row=row_idx, column=col, value=transition.name)
        col += 1
        
        # Target
        ws.cell(row=row_idx, column=col, value=transition.target)
        col += 1
        
        # Level
        if transition.level is not None:
            ws.cell(row=row_idx, column=col, value=transition.level)
        col += 1
        
        # Rule
        ws.cell(row=row_idx, column=col, value=transition.rule)
        col += 1
        
        # Annotations - group by qualifier and combine identifiers
        grouped_annos = {}
        for qualifier, identifier in transition.annotations:
            if qualifier not in grouped_annos:
                grouped_annos[qualifier] = []
            grouped_annos[qualifier].append(_url_to_compact_id(identifier))
        
        for i in range(max_qualifiers):
            if i < len(grouped_annos):
                qualifier = list(grouped_annos.keys())[i]
                identifiers = grouped_annos[qualifier]
                ws.cell(row=row_idx, column=col, value=qualifier)
                ws.cell(row=row_idx, column=col + 1, value=", ".join(identifiers))
            col += 2
        
        # Notes
        for i in range(max_notes):
            if i < len(transition.notes):
                ws.cell(row=row_idx, column=col, value=transition.notes[i])
            col += 1
        
        # Comments
        col += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['E'].width = 40


def _write_interactions_sheet(wb: openpyxl.Workbook, model: InMemoryModel, position: int = 0):
    """Write Interactions sheet"""
    if not model.interactions:
        return  # Skip if no interactions
    
    ws = wb.create_sheet("Interactions", position)
    
    # Header style
    header_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    header_font = Font(bold=True)
    
    # Group annotations by qualifier and determine max unique qualifiers
    max_qualifiers = 0
    max_notes = 0
    for interaction in model.interactions:
        # Group by qualifier
        grouped = {}
        for qualifier, identifier in interaction.annotations:
            if qualifier not in grouped:
                grouped[qualifier] = []
            grouped[qualifier].append(identifier)
        max_qualifiers = max(max_qualifiers, len(grouped))
        max_notes = max(max_notes, len(interaction.notes))
    
    # Build headers
    headers = ["Target", "Source", "Sign"]
    
    # Add Relation/Identifier pairs
    for i in range(max_qualifiers):
        headers.extend([f"Relation{i + 1}", f"Identifier{i + 1}"])
    
    # Add Notes
    for i in range(max_notes):
        headers.append(f"Notes{i + 1}")
    
    headers.append("Comments")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Write interaction data
    for row_idx, interaction in enumerate(model.interactions, 2):
        col = 1
        
        # Target
        ws.cell(row=row_idx, column=col, value=interaction.target)
        col += 1
        
        # Source
        ws.cell(row=row_idx, column=col, value=interaction.source)
        col += 1
        
        # Sign
        ws.cell(row=row_idx, column=col, value=interaction.sign)
        col += 1
        
        # Annotations - group by qualifier and combine identifiers
        grouped_annos = {}
        for qualifier, identifier in interaction.annotations:
            if qualifier not in grouped_annos:
                grouped_annos[qualifier] = []
            grouped_annos[qualifier].append(_url_to_compact_id(identifier))
        
        for i in range(max_qualifiers):
            if i < len(grouped_annos):
                qualifier = list(grouped_annos.keys())[i]
                identifiers = grouped_annos[qualifier]
                ws.cell(row=row_idx, column=col, value=qualifier)
                ws.cell(row=row_idx, column=col + 1, value=", ".join(identifiers))
            col += 2
        
        # Notes
        for i in range(max_notes):
            if i < len(interaction.notes):
                ws.cell(row=row_idx, column=col, value=interaction.notes[i])
            col += 1
        
        # Comments
        col += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15


def _url_to_compact_id(url: str) -> str:
    """Convert identifiers.org URL to compact ID"""
    if "identifiers.org/" in url:
        # Extract the part after identifiers.org/
        parts = url.split("identifiers.org/")
        if len(parts) > 1:
            return parts[1]
    return url


def _person_to_string(person: Person) -> str:
    """Convert Person object to formatted string"""
    parts = []
    if person.family_name:
        parts.append(person.family_name)
    if person.given_name:
        parts.append(person.given_name)
    if person.organization:
        parts.append(f'"{person.organization}"')
    if person.email:
        parts.append(person.email)
    return ", ".join(parts) if parts else ""