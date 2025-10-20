from __future__ import annotations
import os
from typing import Dict, List, Tuple
from openpyxl import load_workbook
import warnings
from datetime import datetime, timezone

from . import spec
from .types import InMemoryModel, ModelInfo, Person, Species, Transition, InteractionEvidence

def _normalize_header(header: str) -> str:
    return header.strip()

def _row_to_dict(headers: List[str], row_values: List[object]) -> Dict[str, object]:
    return {headers[i]: (row_values[i] if i < len(row_values) else None) for i in range(len(headers))}

def _collect_repeated_columns(row: Dict[str, object], prefix: str) -> List[str]:
    values: List[str] = []
    # Accept exact and numbered suffixes
    for key, value in row.items():
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        if spec.is_repeated_column(key, prefix):
            value_str = str(value).strip()
            # For Notes columns, parse note separators <notes1>...</notes1>
            if 'notes' in prefix.lower():
                import re
                # Look for <notesN>content</notesN> patterns
                pattern = r'<notes(\d+)>\s*(.*?)\s*</notes\1>'
                matches = re.findall(pattern, value_str, re.DOTALL)
                if matches:
                    # Extract content from each match
                    for _, content in matches:
                        content = content.strip()
                        if content:
                            values.append(content)
                else:
                    # No separators found, treat as single note
                    if value_str:
                        values.append(value_str)
            else:
                values.append(value_str)
    return values

def _collect_qualifier_pairs(row: Dict[str, object], relation_prefix: str, identifier_prefix: str, validation_warnings: List[str] = None, context: str = "") -> List[Tuple[str, str]]:
    relations = []
    identifiers = []
    for key, value in row.items():
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        if spec.is_repeated_column(key, relation_prefix):
            relations.append((key, str(value).strip()))
        if spec.is_repeated_column(key, identifier_prefix):
            identifiers.append((key, str(value).strip()))
    # Sort by numeric suffix to align pairs
    def sort_key(item: Tuple[str, str]) -> int:
        name = item[0]
        suffix = name[len(name.rstrip("0123456789")):]  # fallback
        # Better: strip prefix explicitly
        return int(''.join(ch for ch in name if ch.isdigit()) or 0)

    relations.sort(key=sort_key)
    identifiers.sort(key=sort_key)
    pairs: List[Tuple[str, str]] = []
    # Default qualifier when relation omitted
    default_species_rel = "is"
    default_trans_inter_rel = "isDescribedBy"
    # Try to pair by index
    if relation_prefix == spec.SPECIES_RELATION_PREFIX:
        default_rel = default_species_rel
    else:
        default_rel = default_trans_inter_rel
    max_len = max(len(relations), len(identifiers))
    for i in range(max_len):
        rel = relations[i][1] if i < len(relations) else default_rel
        # Normalize relation to correct case
        normalized_rel = spec.normalize_relation(rel)
        if normalized_rel is None and validation_warnings is not None:
            validation_warnings.append(f"{context}: Invalid Relation '{rel}'. Valid values: {', '.join(spec.RELATIONS)}")
            normalized_rel = rel  # Keep original if invalid for error tracking
        else:
            normalized_rel = normalized_rel or rel
        
        ident = identifiers[i][1] if i < len(identifiers) else None
        if ident:
            # Split comma-separated identifiers
            for id_part in ident.split(','):
                id_part = id_part.strip()
                if id_part:
                    pairs.append((normalized_rel, id_part))
    return pairs

def _parse_person_string(person_str: str) -> List[str]:
    """Parse person string in format: family_name, given_name, "organization", email"""
    parts = []
    current_part = ""
    in_quotes = False
    i = 0
    
    while i < len(person_str):
        char = person_str[i]
        
        if char == '"':
            if in_quotes:
                in_quotes = False
            else:
                in_quotes = True
        elif char == ',' and not in_quotes:
            parts.append(current_part.strip())
            current_part = ""
        else:
            current_part += char
        i += 1
    
    if current_part.strip():
        parts.append(current_part.strip())
    
    return parts

def read_spreadsheet_to_model(xlsx_path: str) -> tuple[InMemoryModel, list[str]]:
    """Read spreadsheet and return model with validation warnings
    
    Returns:
        tuple: (InMemoryModel, list of warning messages)
    """
    warnings.filterwarnings("ignore", category=UserWarning)
    validation_warnings = []
    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    sheetnames = wb.sheetnames
    # Model sheet (vertical: headers in first column, values in second)
    if spec.SHEET_MODEL in sheetnames:
        ws_model = wb[spec.SHEET_MODEL]
        model_kv: Dict[str, str] = {}
        for r in ws_model.iter_rows(min_row=1, values_only=True):
            if not r:
                continue
            key = _normalize_header(str(r[0] or ""))
            if not key:
                continue
            value = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
            model_kv[key] = value
        model_id = (model_kv.get(spec.MODEL_ID) or "").strip()
        if not model_id:
            model_id = os.path.splitext(os.path.basename(xlsx_path))[0]
        model_info = ModelInfo(
            model_id=model_id,
            name=(model_kv.get(spec.MODEL_NAME) or "").strip() or None,
        )
        # Lists: comma-separated
        def _extend_csv(dst_attr: str, key: str) -> None:
            v = model_kv.get(key)
            if v:
                getattr(model_info, dst_attr).extend([s.strip() for s in str(v).split(",") if str(s).strip()])

        _extend_csv("source_urls", spec.MODEL_SOURCE)
        _extend_csv("described_by", spec.MODEL_PUBLICATION)
        _extend_csv("derived_from", spec.MODEL_ORIGIN_PUBLICATION)
        _extend_csv("derived_from", spec.MODEL_ORIGIN_MODEL)
        _extend_csv("taxons", spec.MODEL_TAXON)
        _extend_csv("biological_processes", spec.MODEL_BIOLOGICAL_PROCESS)
        _extend_csv("versions", spec.MODEL_VERSION)

        model_info.created_iso = (model_kv.get(spec.MODEL_CREATED) or "").strip() or None
        model_info.modified_iso = (model_kv.get(spec.MODEL_MODIFIED) or "").strip() or None
        
        # Set current time if Created field is empty
        if not model_info.created_iso:
            model_info.created_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

        # Repeated rows for creators/contributors/notes: Creator, Creator2, ...
        creators: List[Person] = []
        contributors: List[Person] = []
        notes_list: List[str] = []
        for k, v in model_kv.items():
            if spec.is_repeated_column(k, spec.MODEL_CREATOR_PREFIX) and v:
                parts = _parse_person_string(str(v))
                creators.append(Person(
                    family_name=parts[0] if len(parts) > 0 else None,
                    given_name=parts[1] if len(parts) > 1 else None,
                    organization=parts[2] if len(parts) > 2 else None,
                    email=parts[3] if len(parts) > 3 else None,
                ))
            if spec.is_repeated_column(k, spec.MODEL_CONTRIBUTOR_PREFIX) and v:
                parts = _parse_person_string(str(v))
                contributors.append(Person(
                    family_name=parts[0] if len(parts) > 0 else None,
                    given_name=parts[1] if len(parts) > 1 else None,
                    organization=parts[2] if len(parts) > 2 else None,
                    email=parts[3] if len(parts) > 3 else None,
                ))
            if spec.is_repeated_column(k, spec.MODEL_NOTES_PREFIX) and v:
                notes_list.append(str(v))
        model_info.creators = creators
        model_info.contributors = contributors
        model_info.notes = notes_list
    else: # fallback to filename if missing/empty
        validation_warnings.append("No Model sheet found, using filename as model_id")
        model_info = ModelInfo(
            model_id=os.path.splitext(os.path.basename(xlsx_path))[0]
        )
        # Set current time if Created field is empty
        if not model_info.created_iso:
            model_info.created_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    # Species sheet
    if spec.SHEET_SPECIES not in sheetnames:
        raise ValueError(f"Missing required sheet: {spec.SHEET_SPECIES}")
    ws_species = wb[spec.SHEET_SPECIES]
    sp_headers = [_normalize_header(c.value or "") for c in next(ws_species.iter_rows(min_row=1, max_row=1))[0:ws_species.max_column]]
    species: Dict[str, Species] = {}
    species_row_num = 2
    for r in ws_species.iter_rows(min_row=2, values_only=True):
        rowd = _row_to_dict(sp_headers, list(r or ()))
        sid = str(rowd.get(spec.SPECIES_ID) or "").strip()
        if not sid:
            species_row_num += 1
            continue
        
        # Validate and normalize species type if provided
        species_type = str(rowd.get(spec.SPECIES_TYPE) or "").strip()
        normalized_species_type = None
        if species_type:
            normalized_species_type = spec.normalize_type(species_type)
            if normalized_species_type is None:
                validation_warnings.append(f"Species '{sid}' (row {species_row_num}): Invalid Type '{species_type}'. Valid values: {', '.join(spec.TYPES)}")
        
        sp = Species(
            species_id=sid,
            name=str(rowd.get(spec.SPECIES_NAME) or "").strip() or None,
            compartment=(str(rowd.get(spec.SPECIES_COMPARTMENT) or "").strip() or None),
            constant=_to_bool(rowd.get(spec.SPECIES_CONSTANT)),
            initial_level=_to_int(rowd.get(spec.SPECIES_INITIAL_LEVEL)),
            max_level=_to_int(rowd.get(spec.SPECIES_MAX_LEVEL)),
            annotations=_collect_qualifier_pairs(rowd, spec.SPECIES_RELATION_PREFIX, spec.SPECIES_IDENTIFIER_PREFIX, validation_warnings, f"Species '{sid}' (row {species_row_num})"),
            notes=_collect_repeated_columns(rowd, spec.SPECIES_NOTES_PREFIX),
        )
        species[sid] = sp
        species_row_num += 1
    
    validation_warnings.append(f"Found {len(species)} species")

    # Transitions sheet
    if spec.SHEET_TRANSITIONS not in sheetnames:
        raise ValueError(f"Missing required sheet: {spec.SHEET_TRANSITIONS}")
    ws_trans = wb[spec.SHEET_TRANSITIONS]
    tr_headers = [_normalize_header(c.value or "") for c in next(ws_trans.iter_rows(min_row=1, max_row=1))[0:ws_trans.max_column]]
    transitions: List[Transition] = []
    trans_row_num = 2
    for r in ws_trans.iter_rows(min_row=2, values_only=True):
        rowd = _row_to_dict(tr_headers, list(r or ()))
        target = str(rowd.get(spec.TRANSITION_TARGET) or "").strip()
        rule = str(rowd.get(spec.TRANSITION_RULE) or "").strip()
        if not target and not rule:
            trans_row_num += 1
            continue
        if not target:
            validation_warnings.append(f"Transition row {trans_row_num}: Missing required Target field")
            trans_row_num += 1
            continue
        if not rule:
            validation_warnings.append(f"Transition row {trans_row_num} (Target: {target}): Missing required Rule field")
            trans_row_num += 1
            continue
        
        transitions.append(
            Transition(
                transition_id=str(rowd.get(spec.TRANSITION_ID) or "").strip() or None,
                name=str(rowd.get(spec.TRANSITION_NAME) or "").strip() or None,
                target=target,
                level=_to_int(rowd.get(spec.TRANSITION_LEVEL)),
                rule=rule,
                annotations=_collect_qualifier_pairs(rowd, spec.TRANSITION_RELATION_PREFIX, spec.TRANSITION_IDENTIFIER_PREFIX, validation_warnings, f"Transition '{target}' (row {trans_row_num})"),
                notes=_collect_repeated_columns(rowd, spec.TRANSITION_NOTES_PREFIX),
            )
        )
        trans_row_num += 1
    
    validation_warnings.append(f"Found {len(transitions)} transitions")

    # Interactions sheet: optional
    interactions: List[InteractionEvidence] = []
    if spec.SHEET_INTERACTIONS in sheetnames:
        ws_inter = wb[spec.SHEET_INTERACTIONS]
        in_headers = [_normalize_header(c.value or "") for c in next(ws_inter.iter_rows(min_row=1, max_row=1))[0:ws_inter.max_column]]
        inter_row_num = 2
        for r in ws_inter.iter_rows(min_row=2, values_only=True):
            rowd = _row_to_dict(in_headers, list(r or ()))
            target = str(rowd.get(spec.INTER_TARGET) or "").strip()
            source = str(rowd.get(spec.INTER_SOURCE) or "").strip()
            if not target and not source:
                inter_row_num += 1
                continue
            if not target or not source:
                validation_warnings.append(f"Interaction row {inter_row_num}: Must have both Target and Source")
                inter_row_num += 1
                continue
            
            # Validate and normalize sign
            sign = str(rowd.get(spec.INTER_SIGN) or "").strip() or None
            normalized_sign = None
            if sign:
                normalized_sign = spec.normalize_sign(sign)
                if normalized_sign is None:
                    validation_warnings.append(f"Interaction '{source}→{target}' (row {inter_row_num}): Invalid Sign '{sign}'. Valid values: {', '.join(spec.SIGN)}")
            
            interactions.append(
                InteractionEvidence(
                    target=target,
                    source=source,
                    sign=normalized_sign,
                    annotations=_collect_qualifier_pairs(rowd, spec.INTER_RELATION_PREFIX, spec.INTER_IDENTIFIER_PREFIX, validation_warnings, f"Interaction '{source}→{target}' (row {inter_row_num})"),
                    notes=_collect_repeated_columns(rowd, spec.INTER_NOTES_PREFIX),
                )
            )
            inter_row_num += 1
        validation_warnings.append(f"Found {len(interactions)} interactions")
    else:
        validation_warnings.append("No Interactions sheet found (optional)")

    wb.close()
    
    return InMemoryModel(model=model_info, species=species, transitions=transitions, interactions=interactions), validation_warnings


def _to_bool(v) -> bool | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "t", "1", "yes", "y"):
        return True
    if s in ("false", "f", "0", "no", "n"):
        return False
    return None


def _to_int(v) -> int | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(v)
    except Exception:
        raise ValueError(f"Expected integer, got: {v}")