from __future__ import annotations
import os
import csv
from typing import Dict, List, Tuple
from pathlib import Path
from openpyxl import load_workbook
import warnings
from datetime import datetime, timezone

from . import spec
from .types import QualModel, ModelInfo, Person, Species, Transition, InteractionEvidence
from .tools import validate_and_clean_sid, clean_sid, is_valid_sid


def _make_unique_id(base_id: str, existing_ids: set, id_type: str, context: str, warnings_list: list) -> str:
    """Generate a unique ID by adding suffix if needed.
    
    Args:
        base_id: The original ID
        existing_ids: Set of IDs already used
        id_type: Type of ID for warning message (e.g., "Species_ID", "Transitions_ID")
        context: Context for warning message (e.g., "row 5")
        warnings_list: List to append warnings to
        
    Returns:
        A unique ID (original or with suffix)
    """
    if base_id not in existing_ids:
        return base_id
    
    # Find next available suffix
    suffix = 1
    new_id = f"{base_id}_{suffix}"
    while new_id in existing_ids:
        suffix += 1
        new_id = f"{base_id}_{suffix}"
    
    warnings_list.append(f"{id_type} ({context}): Duplicate ID '{base_id}' renamed to '{new_id}'")
    return new_id


def _generate_sid_from_name(name: str) -> tuple[str, list[str]]:
    """Generate a Species_ID from Name if missing.
    
    Args:
        name: The species name
        
    Returns:
        tuple: (generated_id, warnings_list)
    """
    if not name:
        return "", ["Cannot generate Species_ID from empty Name"]
    
    # Clean the name to make it a valid SId
    generated_id, warnings = clean_sid(name)
    if warnings:
        warnings.insert(0, f"Generated Species_ID from Name '{name}': {warnings[0]}")
    return generated_id, warnings


def _prepare_name_for_rule(name: str, is_valid_sid: bool, is_unique: bool, name_occurrence: int = 0) -> str:
    """Prepare a species name for use in transition rules/interactions.
    
    Args:
        name: The species name
        is_valid_sid: Whether the name conforms to SId format
        is_unique: Whether the name is unique
        name_occurrence: Occurrence number (0 for first/unique, 1 for first duplicate, 2 for second duplicate, etc.)
        
    Returns:
        The name, possibly with suffix and/or quotes if needed
    """
    # Add suffix for duplicates (occurrence 1, 2, etc. get _1, _2 suffixes)
    if not is_unique and name_occurrence > 0:
        suffix_num = name_occurrence
        name_with_suffix = f"{name}_{suffix_num}"
        # Check if the suffixed name is valid SId
        from .tools import is_valid_sid as check_valid_sid
        if check_valid_sid(name_with_suffix):
            return name_with_suffix
        # Need quotes if suffixed name is not valid SId
        return f'"{name_with_suffix}"'
    
    # For unique names or first occurrence
    if is_valid_sid and is_unique:
        return name
    # Need quotes if not valid SId
    return f'"{name}"'


def _resolve_id_to_name(id_or_name: str, species: Dict[str, Species], name_to_id: Dict[str, str], name_counts: Dict[str, int], name_occurrence_map: Dict[str, Dict[int, str]] = None) -> str:
    """Resolve an ID or name to a name (with suffix and/or quotes if needed) for use in rules/interactions.
    
    Args:
        id_or_name: Either a species ID or name from the spreadsheet
        species: Dict of species by ID
        name_to_id: Mapping from name to ID (for use_name=False)
        name_counts: Count of each name occurrence (for uniqueness check)
        name_occurrence_map: Mapping from (name, occurrence) to ID for duplicates
        
    Returns:
        The name, possibly with suffix and/or quotes if needed
    """
    name_occurrence_map = name_occurrence_map or {}
    
    # First try as ID
    if id_or_name in species:
        sp = species[id_or_name]
        if sp.name:
            name = sp.name
            is_valid = is_valid_sid(name)
            is_unique = name_counts.get(name, 0) <= 1
            
            # Find which occurrence this ID corresponds to (0-indexed: 0=first, 1=first duplicate, etc.)
            occurrence = 0
            if name in name_occurrence_map:
                for occ, mapped_sid in name_occurrence_map[name].items():
                    if mapped_sid == id_or_name:
                        occurrence = occ
                        break
            
            return _prepare_name_for_rule(name, is_valid, is_unique, occurrence)
        # Fallback to ID if no name
        return id_or_name
    
    # Try as name (reverse lookup)
    name_to_check = id_or_name.strip('"\'')
    if name_to_check in name_to_id:
        name = name_to_check
        is_valid = is_valid_sid(name)
        is_unique = name_counts.get(name, 0) <= 1
        # For direct name lookup, use first occurrence (0)
        return _prepare_name_for_rule(name, is_valid, is_unique, 0)
    # Also try with original (in case it's stored with quotes somehow)
    if id_or_name in name_to_id:
        name = id_or_name
        is_valid = is_valid_sid(name)
        is_unique = name_counts.get(name, 0) <= 1
        # For direct name lookup, use first occurrence (0)
        return _prepare_name_for_rule(name, is_valid, is_unique, 0)
    
    # Not found, return as-is (will be handled as error later)
    return id_or_name


def _try_resolve_by_id(ref: str, species: Dict[str, Species]) -> bool:
    """Check if a reference can be resolved as a Species ID."""
    # Strip quotes if present
    ref_clean = ref.strip('"\'')
    return ref_clean in species or ref in species


def _try_resolve_by_name(ref: str, name_to_id: Dict[str, str]) -> bool:
    """Check if a reference can be resolved as a Species Name."""
    # Strip quotes if present
    ref_clean = ref.strip('"\'')
    return ref_clean in name_to_id or ref in name_to_id


def _clean_and_try_resolve(ref: str, species: Dict[str, Species], name_to_id: Dict[str, str], 
                          validation_warnings: list, context: str) -> tuple[str, bool, bool]:
    """Try to resolve a reference by cleaning it first (for invalid IDs like 'GeneA/GeneB').
    
    Returns:
        (cleaned_ref, found_as_id, found_as_name) - the cleaned reference and whether it was found
    """
    from .tools import clean_sid
    
    # Strip quotes if present
    ref_clean = ref.strip('"\'')
    
    # Check if it's already valid - if so, no need to clean
    from .tools import is_valid_sid
    if is_valid_sid(ref_clean):
        return ref_clean, False, False
    
    # Try cleaning the reference
    cleaned_ref, clean_warnings = clean_sid(ref_clean)
    if cleaned_ref != ref_clean:
        # Try to find the cleaned version
        found_as_id = cleaned_ref in species
        found_as_name = cleaned_ref in name_to_id
        
        if found_as_id or found_as_name:
            validation_warnings.append(f"{context}: Invalid reference '{ref_clean}' cleaned to '{cleaned_ref}' and found.")
            return cleaned_ref, found_as_id, found_as_name
    
    return ref_clean, False, False


def _resolve_with_fallback(ref: str, species: Dict[str, Species], name_to_id: Dict[str, str], 
                           name_counts: Dict[str, int], name_occurrence_map: Dict[str, Dict[int, str]], 
                           use_name: bool, validation_warnings: list, context: str, 
                           warn_on_switch: bool = True) -> tuple[str, bool]:
    """Resolve a reference with automatic fallback if the current mode doesn't work.
    
    Args:
        warn_on_switch: If False, suppress warnings about mode switching (used when already warned)
    
    Returns:
        (resolved_reference, actual_use_name) - the resolved reference and whether names were actually used
    """
    # Strip quotes if present (quoted names from spreadsheet)
    ref_clean = ref.strip('"\'')
    has_quotes = ref != ref_clean
    
    # Try current mode first (use cleaned reference for checking, but preserve quotes in result if needed)
    if use_name:
        if _try_resolve_by_id(ref_clean, species):
            # User has --use-name but reference is an ID - use ID mode
            if _try_resolve_by_name(ref_clean, name_to_id):
                # Can resolve both ways - use ID mode (prefer ID when ambiguous)
                if warn_on_switch:
                    validation_warnings.append(f"{context}: Reference '{ref_clean}' works with both IDs and Names. Using ID mode (disable --use-name flag).")
                return ref_clean, False
            else:
                # Can only resolve as ID
                if warn_on_switch:
                    validation_warnings.append(f"{context}: Reference '{ref_clean}' is an ID. Using ID mode (disable --use-name flag).")
                return ref_clean, False
        elif _try_resolve_by_name(ref_clean, name_to_id):
            # Can resolve as name
            resolved = _resolve_id_to_name(ref_clean, species, name_to_id, name_counts, name_occurrence_map)
            return resolved, True
        else:
            # Try cleaning the reference if it's invalid
            cleaned_ref, found_as_id, found_as_name = _clean_and_try_resolve(
                ref, species, name_to_id, validation_warnings, context
            )
            if found_as_id:
                if warn_on_switch:
                    validation_warnings.append(f"{context}: Reference '{ref_clean}' is an ID. Using ID mode (disable --use-name flag).")
                return cleaned_ref, False
            elif found_as_name:
                resolved = _resolve_id_to_name(cleaned_ref, species, name_to_id, name_counts, name_occurrence_map)
                return resolved, True
            else:
                # Cannot resolve either way - this is an error
                error_msg = f"{context}: Species reference '{ref}' not found."
                raise ValueError(error_msg)
    else:
        if _try_resolve_by_id(ref_clean, species):
            # Can resolve as ID
            return ref_clean, False
        elif _try_resolve_by_name(ref_clean, name_to_id):
            # User doesn't have --use-name but reference is a name - use name mode
            resolved = _resolve_id_to_name(ref_clean, species, name_to_id, name_counts, name_occurrence_map)
            if warn_on_switch:
                validation_warnings.append(f"{context}: Reference '{ref_clean}' is a Name. Using Name mode (enable --use-name flag).")
            return resolved, True
        else:
            # Try cleaning the reference if it's invalid
            cleaned_ref, found_as_id, found_as_name = _clean_and_try_resolve(
                ref, species, name_to_id, validation_warnings, context
            )
            if found_as_id:
                return cleaned_ref, False
            elif found_as_name:
                resolved = _resolve_id_to_name(cleaned_ref, species, name_to_id, name_counts, name_occurrence_map)
                if warn_on_switch:
                    validation_warnings.append(f"{context}: Reference '{cleaned_ref}' is a Name. Using Name mode (enable --use-name flag).")
                return resolved, True
            else:
                # Cannot resolve either way - this is an error
                error_msg = f"{context}: Species reference '{ref}' not found."
                raise ValueError(error_msg)


def _resolve_rule_with_fallback(rule: str, species: Dict[str, Species], name_to_id: Dict[str, str],
                                name_counts: Dict[str, int], name_occurrence_map: Dict[str, Dict[int, str]],
                                use_name: bool, validation_warnings: list, context: str) -> tuple[str, bool]:
    """Resolve all species references in a rule with automatic fallback.
    
    Returns:
        (resolved_rule, actual_use_name) - the resolved rule and whether names were actually used
    """
    import re
    actual_use_name = use_name
    result = rule
    warned_once = False  # Only warn once about mode switch
    
    # Find all potential species references in the rule
    quoted_pattern = r'"([^"]+)"'
    # Valid SId pattern
    valid_sid_pattern = r'\b([A-Za-z_][A-Za-z0-9_]*)\b'
    # Invalid ID pattern (contains special chars but starts with letter/underscore)
    invalid_id_pattern = r'\b([A-Za-z_][A-Za-z0-9_/\\:\-\.\s]+?)(?=\s|\)|\(|,|$|\band\b|\bor\b|\bnot\b)'
    
    # Find all matches (quoted and unquoted)
    matches = []
    # Find quoted strings first
    for match in re.finditer(quoted_pattern, rule):
        matches.append((match.start(), match.end(), match.group(1), True))  # (start, end, ref, is_quoted)
    
    # Find invalid IDs (with special chars)
    invalid_matches = []
    for match in re.finditer(invalid_id_pattern, rule):
        # Check if this match is inside a quoted string
        in_quotes = any(start <= match.start() < end for start, end, _, _ in matches)
        if not in_quotes:
            # Check if it's not a valid SId
            ref = match.group(1).strip()
            # Skip if it's an operator keyword
            if ref.lower() in ['and', 'or', 'not']:
                continue
            from .tools import is_valid_sid
            if not is_valid_sid(ref):
                invalid_matches.append((match.start(), match.end(), ref, False))
    
    # Find valid SId identifiers (but skip those already in quotes or invalid IDs)
    for match in re.finditer(valid_sid_pattern, rule):
        # Check if this match is inside a quoted string
        in_quotes = any(start <= match.start() < end for start, end, _, _ in matches)
        # Check if this match is inside an invalid ID match
        in_invalid = any(start <= match.start() < end for start, end, _, _ in invalid_matches)
        if not in_quotes and not in_invalid:
            matches.append((match.start(), match.end(), match.group(1), False))
    
    # Add invalid matches
    matches.extend(invalid_matches)
    
    # First pass: determine the mode needed
    mode_switched = False
    for start, end, ref, is_quoted in matches:
        # Skip operators and keywords
        if ref.lower() in ['and', 'or', 'not']:
            continue
        
        # Try to resolve the reference (including cleaning if needed)
        ref_resolved = False
        ref_is_id = False
        ref_is_name = False
        
        # Try direct resolution first
        if _try_resolve_by_id(ref, species):
            ref_resolved = True
            ref_is_id = True
        elif _try_resolve_by_name(ref, name_to_id):
            ref_resolved = True
            ref_is_name = True
        else:
            # Try cleaning the reference
            cleaned_ref, found_as_id, found_as_name = _clean_and_try_resolve(
                ref, species, name_to_id, [], f"{context} (rule reference '{ref}')"
            )
            if found_as_id:
                ref_resolved = True
                ref_is_id = True
            elif found_as_name:
                ref_resolved = True
                ref_is_name = True
        
        # Check if this reference needs a different mode
        if ref_resolved:
            if actual_use_name:
                if ref_is_id and not ref_is_name:
                    # Reference is an ID only - need to switch to ID mode
                    if not mode_switched and not warned_once:
                        validation_warnings.append(f"{context}: Rule contains Species IDs. Using ID mode (disable --use-name flag).")
                        warned_once = True
                    mode_switched = True
                    actual_use_name = False
            else:
                if ref_is_name and not ref_is_id:
                    # Reference is a name only - need to switch to name mode
                    if not mode_switched and not warned_once:
                        validation_warnings.append(f"{context}: Rule contains Species Names. Using Name mode (enable --use-name flag).")
                        warned_once = True
                    mode_switched = True
                    actual_use_name = True
    
    # Second pass: resolve all references using the determined mode
    # Process in reverse order to preserve positions
    replacements = []  # Store (start, end, replacement) tuples
    for start, end, ref, is_quoted in matches:
        # Skip operators and keywords
        if ref.lower() in ['and', 'or', 'not']:
            continue
        
        # Resolve this reference
        try:
            # If it was quoted, pass the quoted version to preserve context
            ref_to_resolve = f'"{ref}"' if is_quoted else ref
            resolved, _ = _resolve_with_fallback(
                ref_to_resolve, species, name_to_id, name_counts, name_occurrence_map,
                actual_use_name, [], f"{context} (rule reference '{ref}')", warn_on_switch=False
            )
            
            # Store replacement (we'll apply them in reverse order)
            replacements.append((start, end, resolved))
        except ValueError as e:
            # Cannot resolve - re-raise as error
            raise ValueError(str(e))
    
    # Apply replacements in reverse order to preserve positions
    for start, end, resolved in reversed(replacements):
        original_text = result[start:end]
        if resolved != original_text:
            result = result[:start] + resolved + result[end:]
    
    return result, actual_use_name


def _resolve_rule_to_names(rule: str, species: Dict[str, Species], name_to_id: Dict[str, str], name_counts: Dict[str, int], name_occurrence_map: Dict[str, Dict[int, str]] = None) -> str:
    """Resolve species IDs in a rule to names (with quotes if needed).
    
    Args:
        rule: The transition rule string
        species: Dict of species by ID
        name_to_id: Mapping from name to ID (for use_name=False)
        name_counts: Count of each name occurrence (for uniqueness check)
        
    Returns:
        Rule with IDs replaced by names (quoted if needed)
    """
    import re
    # Get all species IDs sorted by length (longest first) for greedy matching
    species_ids = sorted(species.keys(), key=len, reverse=True)
    
    result = rule
    # Replace each species ID with its name
    for sid in species_ids:
        sp = species[sid]
        if not sp.name:
            continue
        
        name = sp.name
        is_valid = is_valid_sid(name)
        is_unique = name_counts.get(name, 0) <= 1
        
        # Find which occurrence this ID corresponds to (0-indexed: 0=first, 1=first duplicate, etc.)
        occurrence = 0
        if name in name_occurrence_map:
            for occ, mapped_sid in name_occurrence_map[name].items():
                if mapped_sid == sid:
                    occurrence = occ
                    break
        
        replacement = _prepare_name_for_rule(name, is_valid, is_unique, occurrence)
        
        # Use word boundaries to avoid partial matches
        # But be careful with special characters - use a pattern that matches the ID as a whole token
        # Pattern: \b for word boundary, but also handle cases like "!A" or "A:2"
        pattern = r'\b' + re.escape(sid) + r'(?=\s|&|\||!|\(|\)|>=|<=|>|<|!=|=|:|$)'
        result = re.sub(pattern, replacement, result)
    
    return result

# Common misspellings for sheet/file names
SHEET_NAME_VARIANTS = {
    'model': ['model', 'models'],
    'species': ['species', 'specie'],
    'transitions': ['transitions', 'transition'],
    'interactions': ['interactions', 'interaction'],
}

def _find_csv_file(base_dir: Path, prefix: str, sheet_name: str) -> tuple[str | None, str | None]:
    """Find a CSV file for a given sheet name, checking for common misspellings.
    
    Args:
        base_dir: Directory to search in
        prefix: Prefix for the CSV files (e.g., 'Example' for Example_Species.csv)
        sheet_name: The canonical sheet name (e.g., 'Species', 'Transitions')
    
    Returns:
        tuple: (found_path, warning_message) - warning_message if misspelling was detected
    """
    sheet_lower = sheet_name.lower()
    variants = SHEET_NAME_VARIANTS.get(sheet_lower, [sheet_lower])
    
    if not base_dir.exists():
        return None, None
    
    # Try each variant
    for variant in variants:
        expected_lower = f"{prefix}_{variant}".lower() if prefix else variant
        
        # Case-insensitive search
        for file in base_dir.iterdir():
            if file.is_file() and file.suffix.lower() == '.csv':
                file_lower = file.stem.lower()
                
                if file_lower == expected_lower:
                    warning = None
                    if variant != sheet_lower:
                        warning = f"Found '{file.name}' (misspelling of '{sheet_name}')"
                    return str(file.resolve()), warning
    
    return None, None


def detect_csv_input(input_path: str) -> tuple[bool, dict[str, str], list[str]]:
    """Detect if input is CSV format and find all CSV files.
    
    Args:
        input_path: Input path - can be a folder, prefix, or single CSV file
    
    Returns:
        tuple: (is_csv, files_dict, warnings_list)
            - is_csv: True if CSV input detected
            - files_dict: {'Model': path, 'Species': path, ...}
            - warnings_list: List of warning messages
    """
    input_p = Path(input_path)
    warnings_list = []
    files_dict = {}
    base_dir = None
    prefix = ""
    
    # Helper to check if prefix-style CSV files exist
    def _check_prefix_files(check_dir: Path, check_prefix: str) -> bool:
        if not check_dir.exists():
            return False
        for f in check_dir.iterdir():
            if f.is_file() and f.suffix.lower() == '.csv':
                if f.stem.lower().startswith(check_prefix.lower() + '_'):
                    return True
        return False
    
    # Helper to detect common prefix from CSV files in a directory
    def _detect_prefix_from_dir(check_dir: Path) -> str:
        """Detect common prefix from CSV files in directory."""
        if not check_dir.exists():
            return ""
        csv_files = [f for f in check_dir.iterdir() if f.is_file() and f.suffix.lower() == '.csv']
        if not csv_files:
            return ""
        
        # Try to find common prefix pattern: {prefix}_{sheet}.csv
        prefixes = set()
        for f in csv_files:
            stem = f.stem
            if '_' in stem:
                potential_prefix = stem.rsplit('_', 1)[0]
                # Check if this prefix matches multiple expected sheet names
                for sheet in ['model', 'species', 'transitions', 'interactions']:
                    if stem.lower() == f"{potential_prefix.lower()}_{sheet}":
                        prefixes.add(potential_prefix)
                        break
        
        # If we found a consistent prefix, return it
        if len(prefixes) == 1:
            return list(prefixes)[0]
        return ""
    
    # Check if it's a single CSV file
    if input_p.is_file() and input_p.suffix.lower() == '.csv':
        # Single CSV file - extract prefix from filename (e.g., Example_Species.csv -> Example)
        base_dir = input_p.parent.resolve()
        stem = input_p.stem
        if '_' in stem:
            prefix = stem.rsplit('_', 1)[0]
        else:
            prefix = stem
    else:
        # Determine the parent directory and potential prefix
        if input_p.is_absolute():
            parent_dir = input_p.parent
            potential_prefix = input_p.name
        else:
            parent_dir = Path.cwd() / input_p.parent if str(input_p.parent) != '.' else Path.cwd()
            potential_prefix = input_p.name
        
        # Check if it's a directory first
        if input_p.is_dir():
            # It's a directory - try to detect prefix from CSV files inside
            base_dir = input_p.resolve()
            detected_prefix = _detect_prefix_from_dir(base_dir)
            if detected_prefix:
                prefix = detected_prefix
            else:
                # No prefix detected, look for unprefixed files
                prefix = ""
        # First, check if prefix-style files exist in the parent directory
        # This handles the case where "test" could be a prefix (not a directory)
        elif _check_prefix_files(parent_dir, potential_prefix):
            base_dir = parent_dir
            prefix = potential_prefix
        elif not input_p.exists():
            # Path doesn't exist - could be a prefix, check parent directory
            detected_prefix = _detect_prefix_from_dir(parent_dir)
            if detected_prefix and detected_prefix.lower() == potential_prefix.lower():
                base_dir = parent_dir
                prefix = detected_prefix
            else:
                return False, {}, []
        else:
            return False, {}, []
    
    # Find all sheet files
    for sheet_name in ['Model', 'Species', 'Transitions', 'Interactions']:
        file_path, warning = _find_csv_file(base_dir, prefix, sheet_name)
        if file_path:
            files_dict[sheet_name] = file_path
            if warning:
                warnings_list.append(warning)
    
    # Check required sheets
    if 'Species' not in files_dict:
        warnings_list.append("Missing required file: Species CSV not found")
    if 'Transitions' not in files_dict:
        warnings_list.append("Missing required file: Transitions CSV not found")
    
    return True, files_dict, warnings_list

def _normalize_header(header: str) -> str:
    return header.strip()

def _row_to_dict(headers: List[str], row_values: List[object]) -> Dict[str, object]:
    return {headers[i]: (row_values[i] if i < len(row_values) else None) for i in range(len(headers))}

def _collect_repeated_columns(row: Dict[str, object], prefix: str) -> List[str]:
    values: List[str] = []
    # Accept exact and numbered suffixes
    for key, value in row.items():
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        if spec.is_repeated_column(key, prefix):
            value_str = str(value).strip()
            # For Notes columns, parse note separators <notes1>...</notes1>
            if 'notes' in prefix.lower():
                import re
                # Look for <notesN>content</notesN> patterns
                pattern = r'<notes(\d+)>\s*(.*?)\s*</notes\1>'
                matches = re.findall(pattern, value_str, re.DOTALL)
                if matches:
                    # Extract content from each match
                    for _, content in matches:
                        content = content.strip()
                        if content:
                            values.append(content)
                else:
                    # No separators found, treat as single note
                    if value_str:
                        values.append(value_str)
            else:
                values.append(value_str)
    return values

def _collect_qualifier_pairs(row: Dict[str, object], relation_prefix: str, identifier_prefix: str, validation_warnings: List[str] = None, context: str = "") -> List[Tuple[str, str]]:
    relations = []
    identifiers = []
    for key, value in row.items():
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        if spec.is_repeated_column(key, relation_prefix):
            relations.append((key, str(value).strip()))
        if spec.is_repeated_column(key, identifier_prefix):
            identifiers.append((key, str(value).strip()))
    # Sort by numeric suffix to align pairs
    def sort_key(item: Tuple[str, str]) -> int:
        name = item[0]
        suffix = name[len(name.rstrip("0123456789")):]  # fallback
        # Better: strip prefix explicitly
        return int(''.join(ch for ch in name if ch.isdigit()) or 0)

    relations.sort(key=sort_key)
    identifiers.sort(key=sort_key)
    pairs: List[Tuple[str, str]] = []
    # Default qualifier when relation omitted
    default_species_rel = "is"
    default_trans_inter_rel = "isDescribedBy"
    # Try to pair by index
    if relation_prefix == spec.SPECIES_RELATION_PREFIX:
        default_rel = default_species_rel
    else:
        default_rel = default_trans_inter_rel
    max_len = max(len(relations), len(identifiers))
    for i in range(max_len):
        rel = relations[i][1] if i < len(relations) else default_rel
        # Normalize relation to correct case
        normalized_rel = spec.normalize_relation(rel)
        if normalized_rel is None and validation_warnings is not None:
            validation_warnings.append(f"{context}: Invalid Relation '{rel}'. Valid values: {', '.join(spec.RELATIONS)}")
            normalized_rel = rel  # Keep original if invalid for error tracking
        else:
            normalized_rel = normalized_rel or rel
        
        ident = identifiers[i][1] if i < len(identifiers) else None
        if ident:
            # Split comma-separated identifiers
            for id_part in ident.split(','):
                id_part = id_part.strip()
                if id_part:
                    pairs.append((normalized_rel, id_part))
    return pairs

def _parse_person_string(person_str: str) -> List[str]:
    """Parse person string in format: family_name, given_name, "organization", email"""
    parts = []
    current_part = ""
    in_quotes = False
    i = 0
    
    while i < len(person_str):
        char = person_str[i]
        
        if char == '"':
            if in_quotes:
                in_quotes = False
            else:
                in_quotes = True
        elif char == ',' and not in_quotes:
            parts.append(current_part.strip())
            current_part = ""
        else:
            current_part += char
        i += 1
    
    if current_part.strip():
        parts.append(current_part.strip())
    
    return parts

def read_spreadsheet_to_model(xlsx_path: str, use_name: bool = False) -> tuple[QualModel, list[str]]:
    """Read spreadsheet and return model with validation warnings
    
    Args:
        xlsx_path: Path to XLSX file
        use_name: If True, use Species_ID in rules/interactions. If False, use Name (with quotes if needed).
    
    Returns:
        tuple: (QualModel, list of warning messages)
    """
    warnings.filterwarnings("ignore", category=UserWarning)
    validation_warnings = []
    wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    sheetnames = wb.sheetnames
    # Model sheet (vertical: headers in first column, values in second)
    if spec.SHEET_MODEL in sheetnames:
        ws_model = wb[spec.SHEET_MODEL]
        model_kv: Dict[str, str] = {}
        for r in ws_model.iter_rows(min_row=1, values_only=True):
            if not r:
                continue
            key = _normalize_header(str(r[0] or ""))
            if not key:
                continue
            value = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
            model_kv[key] = value
        model_id = (model_kv.get(spec.MODEL_ID) or "").strip()
        if not model_id:
            model_id = os.path.splitext(os.path.basename(xlsx_path))[0]
        
        # Validate and clean Model_ID (SId format)
        model_id, model_id_warnings = validate_and_clean_sid(model_id, "Model_ID", "Model sheet")
        validation_warnings.extend(model_id_warnings)
        
        model_info = ModelInfo(
            model_id=model_id,
            name=(model_kv.get(spec.MODEL_NAME) or "").strip() or None,
        )
        # Lists: comma-separated
        def _extend_csv(dst_attr: str, key: str) -> None:
            v = model_kv.get(key)
            if v:
                getattr(model_info, dst_attr).extend([s.strip() for s in str(v).split(",") if str(s).strip()])

        _extend_csv("source_urls", spec.MODEL_SOURCE)
        _extend_csv("described_by", spec.MODEL_PUBLICATION)
        _extend_csv("derived_from", spec.MODEL_ORIGIN_PUBLICATION)
        _extend_csv("derived_from", spec.MODEL_ORIGIN_MODEL)
        _extend_csv("taxons", spec.MODEL_TAXON)
        _extend_csv("biological_processes", spec.MODEL_BIOLOGICAL_PROCESS)
        _extend_csv("versions", spec.MODEL_VERSION)

        model_info.created_iso = (model_kv.get(spec.MODEL_CREATED) or "").strip() or None
        model_info.modified_iso = (model_kv.get(spec.MODEL_MODIFIED) or "").strip() or None
        
        # Set current time if Created field is empty
        if not model_info.created_iso:
            model_info.created_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

        # Repeated rows for creators/contributors/notes: Creator, Creator2, ...
        creators: List[Person] = []
        contributors: List[Person] = []
        notes_list: List[str] = []
        for k, v in model_kv.items():
            if spec.is_repeated_column(k, spec.MODEL_CREATOR_PREFIX) and v:
                parts = _parse_person_string(str(v))
                creators.append(Person(
                    family_name=parts[0] if len(parts) > 0 else None,
                    given_name=parts[1] if len(parts) > 1 else None,
                    organization=parts[2] if len(parts) > 2 else None,
                    email=parts[3] if len(parts) > 3 else None,
                ))
            if spec.is_repeated_column(k, spec.MODEL_CONTRIBUTOR_PREFIX) and v:
                parts = _parse_person_string(str(v))
                contributors.append(Person(
                    family_name=parts[0] if len(parts) > 0 else None,
                    given_name=parts[1] if len(parts) > 1 else None,
                    organization=parts[2] if len(parts) > 2 else None,
                    email=parts[3] if len(parts) > 3 else None,
                ))
            if spec.is_repeated_column(k, spec.MODEL_NOTES_PREFIX) and v:
                notes_list.append(str(v))
        model_info.creators = creators
        model_info.contributors = contributors
        model_info.notes = notes_list
    else: # fallback to filename if missing/empty
        validation_warnings.append("No Model sheet found, using filename as model_id")
        model_id = os.path.splitext(os.path.basename(xlsx_path))[0]
        
        # Validate and clean Model_ID (SId format)
        model_id, model_id_warnings = validate_and_clean_sid(model_id, "Model_ID", "from filename")
        validation_warnings.extend(model_id_warnings)
        
        model_info = ModelInfo(model_id=model_id)
        # Set current time if Created field is empty
        if not model_info.created_iso:
            model_info.created_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    # Species sheet
    if spec.SHEET_SPECIES not in sheetnames:
        raise ValueError(f"Missing required sheet: {spec.SHEET_SPECIES}")
    ws_species = wb[spec.SHEET_SPECIES]
    sp_headers = [_normalize_header(c.value or "") for c in next(ws_species.iter_rows(min_row=1, max_row=1))[0:ws_species.max_column]]
    species: Dict[str, Species] = {}
    used_species_ids: set = set()  # Track all species IDs (including renamed ones)
    name_to_id: Dict[str, str] = {}  # Map name to ID for use_name=True
    used_names: set = set()  # Track name usage for uniqueness when use_name=True
    name_counts: Dict[str, int] = {}  # Count occurrences of each name
    name_occurrence_map: Dict[str, Dict[int, str]] = {}  # Map (name, occurrence) to ID for duplicate names
    species_row_num = 2
    for r in ws_species.iter_rows(min_row=2, values_only=True):
        rowd = _row_to_dict(sp_headers, list(r or ()))
        name = str(rowd.get(spec.SPECIES_NAME) or "").strip() or None
        
        # Get or generate Species_ID
        sid = str(rowd.get(spec.SPECIES_ID) or "").strip()
        
        # Skip completely empty rows (no Species_ID, no Name, and no other data)
        if not sid and not name:
            # Check if row has any other non-empty data
            has_other_data = any(
                str(rowd.get(key) or "").strip() 
                for key in rowd.keys() 
                if key not in (spec.SPECIES_ID, spec.SPECIES_NAME)
            )
            if not has_other_data:
                # Completely empty row - skip silently
                species_row_num += 1
                continue
        
        if not sid:
            # Try to generate from Name if available
            if name:
                generated_id, gen_warnings = _generate_sid_from_name(name)
                validation_warnings.extend(gen_warnings)
                sid = generated_id
            else:
                # Both Species_ID and Name are missing but row has other data - skip with warning
                validation_warnings.append(f"Species row {species_row_num}: Missing both Species_ID and Name - skipping row")
                species_row_num += 1
                continue
        
        # Validate and clean Species_ID (SId format)
        sid, sid_warnings = validate_and_clean_sid(sid, "Species_ID", f"row {species_row_num}")
        validation_warnings.extend(sid_warnings)
        
        # Make unique if duplicate
        sid = _make_unique_id(sid, used_species_ids, "Species_ID", f"row {species_row_num}", validation_warnings)
        used_species_ids.add(sid)
        
        # Track name usage
        if name:
            if name in name_counts:
                name_counts[name] += 1
                occurrence = name_counts[name] - 1  # 0-indexed occurrence
            else:
                name_counts[name] = 1
                occurrence = 0
            used_names.add(name)
            name_to_id[name] = sid  # First occurrence maps directly
            # Track occurrence mapping for duplicates
            if name not in name_occurrence_map:
                name_occurrence_map[name] = {}
            name_occurrence_map[name][occurrence] = sid
        
        # Validate and clean Compartment (SId format) if provided
        compartment = str(rowd.get(spec.SPECIES_COMPARTMENT) or "").strip() or None
        if compartment:
            compartment, comp_warnings = validate_and_clean_sid(compartment, "Compartment", f"Species '{sid}', row {species_row_num}")
            validation_warnings.extend(comp_warnings)
        
        # Validate and normalize species type if provided
        species_type = str(rowd.get(spec.SPECIES_TYPE) or "").strip()
        normalized_species_type = None
        if species_type:
            normalized_species_type = spec.normalize_type(species_type)
            if normalized_species_type is None:
                validation_warnings.append(f"Species '{sid}' (row {species_row_num}): Invalid Type '{species_type}'. Valid values: {', '.join(spec.TYPES)}")
        
        sp = Species(
            species_id=sid,
            name=name,
            compartment=compartment,
            constant=_to_bool(rowd.get(spec.SPECIES_CONSTANT)),
            initial_level=_to_int(rowd.get(spec.SPECIES_INITIAL_LEVEL)),
            max_level=_to_int(rowd.get(spec.SPECIES_MAX_LEVEL)),
            type=normalized_species_type,
            annotations=_collect_qualifier_pairs(rowd, spec.SPECIES_RELATION_PREFIX, spec.SPECIES_IDENTIFIER_PREFIX, validation_warnings, f"Species '{sid}' (row {species_row_num})"),
            notes=_collect_repeated_columns(rowd, spec.SPECIES_NOTES_PREFIX),
        )
        species[sid] = sp
        species_row_num += 1
    
    validation_warnings.append(f"Found {len(species)} species")
    
    # Transitions sheet
    if spec.SHEET_TRANSITIONS not in sheetnames:
        raise ValueError(f"Missing required sheet: {spec.SHEET_TRANSITIONS}")
    ws_trans = wb[spec.SHEET_TRANSITIONS]
    tr_headers = [_normalize_header(c.value or "") for c in next(ws_trans.iter_rows(min_row=1, max_row=1))[0:ws_trans.max_column]]
    transitions: List[Transition] = []
    used_trans_ids: set = set()  # Track all transition IDs (including renamed ones)
    trans_row_num = 2
    for r in ws_trans.iter_rows(min_row=2, values_only=True):
        rowd = _row_to_dict(tr_headers, list(r or ()))
        target = str(rowd.get(spec.TRANSITION_TARGET) or "").strip()
        rule = str(rowd.get(spec.TRANSITION_RULE) or "").strip()
        if not target and not rule:
            trans_row_num += 1
            continue
        if not target:
            validation_warnings.append(f"Transition row {trans_row_num}: Missing required Target field")
            trans_row_num += 1
            continue
        if not rule:
            validation_warnings.append(f"Transition row {trans_row_num} (Target: {target}): Missing required Rule field")
            trans_row_num += 1
            continue
        
        # Resolve target and rule with automatic fallback
        context = f"Transition row {trans_row_num}"
        target, actual_use_name = _resolve_with_fallback(
            target, species, name_to_id, name_counts, name_occurrence_map,
            use_name, validation_warnings, f"{context} (Target: {target})"
        )
        rule, actual_use_name = _resolve_rule_with_fallback(
            rule, species, name_to_id, name_counts, name_occurrence_map,
            actual_use_name, validation_warnings, context
        )
        
        # Validate and clean Transitions_ID (SId format) if provided
        trans_id = str(rowd.get(spec.TRANSITION_ID) or "").strip() or None
        if trans_id:
            trans_id, trans_id_warnings = validate_and_clean_sid(trans_id, "Transitions_ID", f"row {trans_row_num}")
            validation_warnings.extend(trans_id_warnings)
            
            # Make unique if duplicate
            trans_id = _make_unique_id(trans_id, used_trans_ids, "Transitions_ID", f"row {trans_row_num}", validation_warnings)
            used_trans_ids.add(trans_id)
        
        transitions.append(
            Transition(
                transition_id=trans_id,
                name=str(rowd.get(spec.TRANSITION_NAME) or "").strip() or None,
                target=target,
                level=_to_int(rowd.get(spec.TRANSITION_LEVEL)),
                rule=rule,
                annotations=_collect_qualifier_pairs(rowd, spec.TRANSITION_RELATION_PREFIX, spec.TRANSITION_IDENTIFIER_PREFIX, validation_warnings, f"Transition '{target}' (row {trans_row_num})"),
                notes=_collect_repeated_columns(rowd, spec.TRANSITION_NOTES_PREFIX),
            )
        )
        trans_row_num += 1
    
    validation_warnings.append(f"Found {len(transitions)} transitions")

    # Interactions sheet: optional
    interactions: List[InteractionEvidence] = []
    if spec.SHEET_INTERACTIONS in sheetnames:
        ws_inter = wb[spec.SHEET_INTERACTIONS]
        in_headers = [_normalize_header(c.value or "") for c in next(ws_inter.iter_rows(min_row=1, max_row=1))[0:ws_inter.max_column]]
        inter_row_num = 2
        for r in ws_inter.iter_rows(min_row=2, values_only=True):
            rowd = _row_to_dict(in_headers, list(r or ()))
            target = str(rowd.get(spec.INTER_TARGET) or "").strip()
            source = str(rowd.get(spec.INTER_SOURCE) or "").strip()
            if not target and not source:
                inter_row_num += 1
                continue
            if not target or not source:
                validation_warnings.append(f"Interaction row {inter_row_num}: Must have both Target and Source")
                inter_row_num += 1
                continue
            
            # Resolve target and source with automatic fallback
            context = f"Interaction row {inter_row_num}"
            target, actual_use_name = _resolve_with_fallback(
                target, species, name_to_id, name_counts, name_occurrence_map,
                use_name, validation_warnings, f"{context} (Target: {target})"
            )
            source, actual_use_name = _resolve_with_fallback(
                source, species, name_to_id, name_counts, name_occurrence_map,
                actual_use_name, validation_warnings, f"{context} (Source: {source})"
            )
            
            # Validate and normalize sign
            sign = str(rowd.get(spec.INTER_SIGN) or "").strip() or None
            normalized_sign = None
            if sign:
                normalized_sign = spec.normalize_sign(sign)
                if normalized_sign is None:
                    validation_warnings.append(f"Interaction '{source}→{target}' (row {inter_row_num}): Invalid Sign '{sign}'. Valid values: {', '.join(spec.SIGN)}")
            
            interactions.append(
                InteractionEvidence(
                    target=target,
                    source=source,
                    sign=normalized_sign,
                    annotations=_collect_qualifier_pairs(rowd, spec.INTER_RELATION_PREFIX, spec.INTER_IDENTIFIER_PREFIX, validation_warnings, f"Interaction '{source}→{target}' (row {inter_row_num})"),
                    notes=_collect_repeated_columns(rowd, spec.INTER_NOTES_PREFIX),
                )
            )
            inter_row_num += 1
        validation_warnings.append(f"Found {len(interactions)} interactions")
    else:
        validation_warnings.append("No Interactions sheet found (optional)")

    wb.close()
    
    return QualModel(model=model_info, species=species, transitions=transitions, interactions=interactions), validation_warnings


def _to_bool(v) -> bool | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "t", "1", "yes", "y"):
        return True
    if s in ("false", "f", "0", "no", "n"):
        return False
    return None


def _to_int(v) -> int | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(v)
    except Exception:
        raise ValueError(f"Expected integer, got: {v}")


def _read_csv_to_rows(csv_path: str) -> tuple[list[str], list[dict[str, object]]]:
    """Read a CSV file and return headers and rows as dicts.
    
    Returns:
        tuple: (headers, list of row dicts)
    """
    rows = []
    headers = []
    with open(csv_path, 'r', newline='', encoding='utf-8-sig') as f:  # utf-8-sig strips BOM
        reader = csv.reader(f)
        try:
            header_row = next(reader)
            headers = [_normalize_header(h) for h in header_row]
        except StopIteration:
            return [], []
        
        for row in reader:
            row_dict = _row_to_dict(headers, row)
            rows.append(row_dict)
    
    return headers, rows


def read_csv_to_model(csv_files: dict[str, str], use_name: bool = False) -> tuple[QualModel, list[str]]:
    """Read CSV files and return model with validation warnings.
    
    Args:
        csv_files: Dict mapping sheet names to file paths
                   {'Model': path, 'Species': path, 'Transitions': path, 'Interactions': path}
        use_name: If True, use Species_ID in rules/interactions. If False, use Name (with quotes if needed).
    
    Returns:
        tuple: (QualModel, list of warning messages)
    """
    validation_warnings = []
    
    # Model CSV (vertical: headers in first column, values in second)
    if 'Model' in csv_files:
        model_kv: Dict[str, str] = {}
        with open(csv_files['Model'], 'r', newline='', encoding='utf-8-sig') as f:  # utf-8-sig strips BOM
            reader = csv.reader(f)
            for row in reader:
                if not row:
                    continue
                key = _normalize_header(str(row[0] or ""))
                if not key:
                    continue
                value = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                model_kv[key] = value
        
        model_id = (model_kv.get(spec.MODEL_ID) or "").strip()
        if not model_id:
            # Fallback to filename
            model_id = Path(csv_files['Model']).stem.replace('_Model', '').replace('_model', '')
        
        # Validate and clean Model_ID (SId format)
        model_id, model_id_warnings = validate_and_clean_sid(model_id, "Model_ID", "Model CSV")
        validation_warnings.extend(model_id_warnings)
        
        model_info = ModelInfo(
            model_id=model_id,
            name=(model_kv.get(spec.MODEL_NAME) or "").strip() or None,
        )
        
        # Lists: comma-separated
        def _extend_csv_field(dst_attr: str, key: str) -> None:
            v = model_kv.get(key)
            if v:
                getattr(model_info, dst_attr).extend([s.strip() for s in str(v).split(",") if str(s).strip()])
        
        _extend_csv_field("source_urls", spec.MODEL_SOURCE)
        _extend_csv_field("described_by", spec.MODEL_PUBLICATION)
        _extend_csv_field("derived_from", spec.MODEL_ORIGIN_PUBLICATION)
        _extend_csv_field("derived_from", spec.MODEL_ORIGIN_MODEL)
        _extend_csv_field("taxons", spec.MODEL_TAXON)
        _extend_csv_field("biological_processes", spec.MODEL_BIOLOGICAL_PROCESS)
        _extend_csv_field("versions", spec.MODEL_VERSION)
        
        model_info.created_iso = (model_kv.get(spec.MODEL_CREATED) or "").strip() or None
        model_info.modified_iso = (model_kv.get(spec.MODEL_MODIFIED) or "").strip() or None
        
        # Set current time if Created field is empty
        if not model_info.created_iso:
            model_info.created_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        
        # Repeated rows for creators/contributors/notes
        creators: List[Person] = []
        contributors: List[Person] = []
        notes_list: List[str] = []
        for k, v in model_kv.items():
            if spec.is_repeated_column(k, spec.MODEL_CREATOR_PREFIX) and v:
                parts = _parse_person_string(str(v))
                creators.append(Person(
                    family_name=parts[0] if len(parts) > 0 else None,
                    given_name=parts[1] if len(parts) > 1 else None,
                    organization=parts[2] if len(parts) > 2 else None,
                    email=parts[3] if len(parts) > 3 else None,
                ))
            if spec.is_repeated_column(k, spec.MODEL_CONTRIBUTOR_PREFIX) and v:
                parts = _parse_person_string(str(v))
                contributors.append(Person(
                    family_name=parts[0] if len(parts) > 0 else None,
                    given_name=parts[1] if len(parts) > 1 else None,
                    organization=parts[2] if len(parts) > 2 else None,
                    email=parts[3] if len(parts) > 3 else None,
                ))
            if spec.is_repeated_column(k, spec.MODEL_NOTES_PREFIX) and v:
                notes_list.append(str(v))
        model_info.creators = creators
        model_info.contributors = contributors
        model_info.notes = notes_list
    else:
        validation_warnings.append("No Model CSV found, using default model_id")
        # Try to get model_id from species file name
        species_path = csv_files.get('Species', '')
        model_id = Path(species_path).stem.replace('_Species', '').replace('_species', '') if species_path else 'unknown'
        
        # Validate and clean Model_ID (SId format)
        model_id, model_id_warnings = validate_and_clean_sid(model_id, "Model_ID", "from filename")
        validation_warnings.extend(model_id_warnings)
        
        model_info = ModelInfo(model_id=model_id)
        if not model_info.created_iso:
            model_info.created_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    
    # Species CSV
    if 'Species' not in csv_files:
        raise ValueError("Missing required CSV file: Species")
    
    _, species_rows = _read_csv_to_rows(csv_files['Species'])
    species: Dict[str, Species] = {}
    used_species_ids: set = set()  # Track all species IDs (including renamed ones)
    name_to_id: Dict[str, str] = {}  # Map name to ID for use_name=False
    used_names: set = set()  # Track name usage for uniqueness when use_name=False
    name_counts: Dict[str, int] = {}  # Count occurrences of each name
    name_occurrence_map: Dict[str, Dict[int, str]] = {}  # Map (name, occurrence) to ID for duplicate names
    species_row_num = 2
    for rowd in species_rows:
        name = str(rowd.get(spec.SPECIES_NAME) or "").strip() or None
        
        # Get or generate Species_ID
        sid = str(rowd.get(spec.SPECIES_ID) or "").strip()
        
        # Skip completely empty rows (no Species_ID, no Name, and no other data)
        if not sid and not name:
            # Check if row has any other non-empty data
            has_other_data = any(
                str(rowd.get(key) or "").strip() 
                for key in rowd.keys() 
                if key not in (spec.SPECIES_ID, spec.SPECIES_NAME)
            )
            if not has_other_data:
                # Completely empty row - skip silently
                species_row_num += 1
                continue
        
        if not sid:
            # Try to generate from Name if available
            if name:
                generated_id, gen_warnings = _generate_sid_from_name(name)
                validation_warnings.extend(gen_warnings)
                sid = generated_id
            else:
                # Both Species_ID and Name are missing but row has other data - skip with warning
                validation_warnings.append(f"Species row {species_row_num}: Missing both Species_ID and Name - skipping row")
                species_row_num += 1
                continue
        
        # Validate and clean Species_ID (SId format)
        sid, sid_warnings = validate_and_clean_sid(sid, "Species_ID", f"row {species_row_num}")
        validation_warnings.extend(sid_warnings)
        
        # Make unique if duplicate
        sid = _make_unique_id(sid, used_species_ids, "Species_ID", f"row {species_row_num}", validation_warnings)
        used_species_ids.add(sid)
        
        # Track name usage
        if name:
            if name in name_counts:
                name_counts[name] += 1
                occurrence = name_counts[name] - 1  # 0-indexed occurrence
            else:
                name_counts[name] = 1
                occurrence = 0
            used_names.add(name)
            name_to_id[name] = sid  # First occurrence maps directly
            # Track occurrence mapping for duplicates
            if name not in name_occurrence_map:
                name_occurrence_map[name] = {}
            name_occurrence_map[name][occurrence] = sid
        
        # Validate and clean Compartment (SId format) if provided
        compartment = str(rowd.get(spec.SPECIES_COMPARTMENT) or "").strip() or None
        if compartment:
            compartment, comp_warnings = validate_and_clean_sid(compartment, "Compartment", f"Species '{sid}', row {species_row_num}")
            validation_warnings.extend(comp_warnings)
        
        # Validate and normalize species type if provided
        species_type = str(rowd.get(spec.SPECIES_TYPE) or "").strip()
        normalized_species_type = None
        if species_type:
            normalized_species_type = spec.normalize_type(species_type)
            if normalized_species_type is None:
                validation_warnings.append(f"Species '{sid}' (row {species_row_num}): Invalid Type '{species_type}'. Valid values: {', '.join(spec.TYPES)}")
        
        sp = Species(
            species_id=sid,
            name=name,
            compartment=compartment,
            constant=_to_bool(rowd.get(spec.SPECIES_CONSTANT)),
            initial_level=_to_int(rowd.get(spec.SPECIES_INITIAL_LEVEL)),
            max_level=_to_int(rowd.get(spec.SPECIES_MAX_LEVEL)),
            type=normalized_species_type,
            annotations=_collect_qualifier_pairs(rowd, spec.SPECIES_RELATION_PREFIX, spec.SPECIES_IDENTIFIER_PREFIX, validation_warnings, f"Species '{sid}' (row {species_row_num})"),
            notes=_collect_repeated_columns(rowd, spec.SPECIES_NOTES_PREFIX),
        )
        species[sid] = sp
        species_row_num += 1
    
    validation_warnings.append(f"Found {len(species)} species")
    
    # Validate names when use_name=True
    if use_name:
        missing_names = [sid for sid, sp in species.items() if not sp.name]
        if missing_names:
            validation_warnings.append(f"Warning: {len(missing_names)} species missing Names (IDs: {', '.join(missing_names[:5])}{'...' if len(missing_names) > 5 else ''}). Cannot use Names - falling back to IDs.")
            use_name = False  # Fall back to using IDs
    
    # Transitions CSV
    if 'Transitions' not in csv_files:
        raise ValueError("Missing required CSV file: Transitions")
    
    _, trans_rows = _read_csv_to_rows(csv_files['Transitions'])
    transitions: List[Transition] = []
    used_trans_ids: set = set()  # Track all transition IDs (including renamed ones)
    trans_row_num = 2
    for rowd in trans_rows:
        target = str(rowd.get(spec.TRANSITION_TARGET) or "").strip()
        rule = str(rowd.get(spec.TRANSITION_RULE) or "").strip()
        if not target and not rule:
            trans_row_num += 1
            continue
        if not target:
            validation_warnings.append(f"Transition row {trans_row_num}: Missing required Target field")
            trans_row_num += 1
            continue
        if not rule:
            validation_warnings.append(f"Transition row {trans_row_num} (Target: {target}): Missing required Rule field")
            trans_row_num += 1
            continue
        
        # Resolve target and rule with automatic fallback
        context = f"Transition row {trans_row_num}"
        target, actual_use_name = _resolve_with_fallback(
            target, species, name_to_id, name_counts, name_occurrence_map,
            use_name, validation_warnings, f"{context} (Target: {target})"
        )
        rule, actual_use_name = _resolve_rule_with_fallback(
            rule, species, name_to_id, name_counts, name_occurrence_map,
            actual_use_name, validation_warnings, context
        )
        
        # Validate and clean Transitions_ID (SId format) if provided
        trans_id = str(rowd.get(spec.TRANSITION_ID) or "").strip() or None
        if trans_id:
            trans_id, trans_id_warnings = validate_and_clean_sid(trans_id, "Transitions_ID", f"row {trans_row_num}")
            validation_warnings.extend(trans_id_warnings)
            
            # Make unique if duplicate
            trans_id = _make_unique_id(trans_id, used_trans_ids, "Transitions_ID", f"row {trans_row_num}", validation_warnings)
            used_trans_ids.add(trans_id)
        
        transitions.append(
            Transition(
                transition_id=trans_id,
                name=str(rowd.get(spec.TRANSITION_NAME) or "").strip() or None,
                target=target,
                level=_to_int(rowd.get(spec.TRANSITION_LEVEL)),
                rule=rule,
                annotations=_collect_qualifier_pairs(rowd, spec.TRANSITION_RELATION_PREFIX, spec.TRANSITION_IDENTIFIER_PREFIX, validation_warnings, f"Transition '{target}' (row {trans_row_num})"),
                notes=_collect_repeated_columns(rowd, spec.TRANSITION_NOTES_PREFIX),
            )
        )
        trans_row_num += 1
    
    validation_warnings.append(f"Found {len(transitions)} transitions")
    
    # Interactions CSV (optional)
    interactions: List[InteractionEvidence] = []
    if 'Interactions' in csv_files:
        _, inter_rows = _read_csv_to_rows(csv_files['Interactions'])
        inter_row_num = 2
        for rowd in inter_rows:
            target = str(rowd.get(spec.INTER_TARGET) or "").strip()
            source = str(rowd.get(spec.INTER_SOURCE) or "").strip()
            if not target and not source:
                inter_row_num += 1
                continue
            if not target or not source:
                validation_warnings.append(f"Interaction row {inter_row_num}: Must have both Target and Source")
                inter_row_num += 1
                continue
            
            # Resolve target and source with automatic fallback
            context = f"Interaction row {inter_row_num}"
            target, actual_use_name = _resolve_with_fallback(
                target, species, name_to_id, name_counts, name_occurrence_map,
                use_name, validation_warnings, f"{context} (Target: {target})"
            )
            source, actual_use_name = _resolve_with_fallback(
                source, species, name_to_id, name_counts, name_occurrence_map,
                actual_use_name, validation_warnings, f"{context} (Source: {source})"
            )
            
            # Validate and normalize sign
            sign = str(rowd.get(spec.INTER_SIGN) or "").strip() or None
            normalized_sign = None
            if sign:
                normalized_sign = spec.normalize_sign(sign)
                if normalized_sign is None:
                    validation_warnings.append(f"Interaction '{source}→{target}' (row {inter_row_num}): Invalid Sign '{sign}'. Valid values: {', '.join(spec.SIGN)}")
            
            interactions.append(
                InteractionEvidence(
                    target=target,
                    source=source,
                    sign=normalized_sign,
                    annotations=_collect_qualifier_pairs(rowd, spec.INTER_RELATION_PREFIX, spec.INTER_IDENTIFIER_PREFIX, validation_warnings, f"Interaction '{source}→{target}' (row {inter_row_num})"),
                    notes=_collect_repeated_columns(rowd, spec.INTER_NOTES_PREFIX),
                )
            )
            inter_row_num += 1
        validation_warnings.append(f"Found {len(interactions)} interactions")
    else:
        validation_warnings.append("No Interactions CSV found (optional)")
    
    return QualModel(model=model_info, species=species, transitions=transitions, interactions=interactions), validation_warnings