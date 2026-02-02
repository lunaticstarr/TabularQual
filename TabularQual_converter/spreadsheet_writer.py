from __future__ import annotations

from typing import List, Dict, Any
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path
import shutil

from .types import QualModel, Person
from . import spec
from .tools import is_valid_sid


def _build_name_deduplication_map(species: Dict[str, Any]) -> Dict[str, str]:
    """Build a map from (species_id, original_name) to deduplicated name with suffix.
    
    Returns:
        Dict mapping (species_id, original_name) to suffixed name (e.g., "Name_1", "Name_2")
    """
    name_counts = {}
    name_occurrence_map = {}
    name_current_count = {}
    
    # First pass: count name occurrences
    for sp in species.values():
        if sp.name:
            name = sp.name.strip()  # Strip leading/trailing spaces
            name_counts[name] = name_counts.get(name, 0) + 1
    
    # Second pass: build occurrence map
    for sp in sorted(species.values(), key=lambda s: s.species_id):
        if sp.name:
            name = sp.name.strip()
            if name_counts[name] > 1:
                # Duplicate name - need suffix
                occurrence = name_current_count.get(name, 0)
                name_current_count[name] = occurrence + 1
                # First occurrence (0-indexed) gets no suffix, subsequent get _1, _2, etc.
                if occurrence == 0:
                    name_occurrence_map[sp.species_id] = name
                else:
                    name_occurrence_map[sp.species_id] = f"{name}_{occurrence}"
            else:
                # Unique name - no suffix needed
                name_occurrence_map[sp.species_id] = name
    
    return name_occurrence_map


def _resolve_id_to_name_for_output(sid: str, species: Dict[str, Any], use_name: bool, 
                                   name_dedup_map: Dict[str, str] = None) -> str:
    """Resolve a species ID to Name for output when use_name=True.
    
    Args:
        sid: Species ID
        species: Dict of species by ID
        use_name: If True, return Name (with quotes if needed). If False, return ID.
        name_dedup_map: Optional map from species_id to deduplicated name (with suffix if needed)
        
    Returns:
        Name (possibly quoted) if use_name=True, or ID if use_name=False
    """
    if not use_name:
        return sid
    
    if sid not in species:
        return sid  # Not found, return as-is
    
    sp = species[sid]
    if not sp.name:
        return sid  # No name, use ID
    
    # Use deduplicated name if map is provided
    if name_dedup_map and sid in name_dedup_map:
        name = name_dedup_map[sid]
    else:
        name = sp.name.strip()
    
    # Check if name is valid SId
    is_valid = is_valid_sid(name)
    
    if is_valid:
        return name
    # Need quotes if not valid SId
    return f'"{name}"'


def _resolve_rule_ids_to_names(rule: str, species: Dict[str, Any], use_name: bool,
                               name_dedup_map: Dict[str, str] = None) -> str:
    """Resolve species IDs in a rule to Names when use_name=True.
    
    Args:
        rule: Transition rule string
        species: Dict of species by ID
        use_name: If True, replace IDs with Names. If False, return rule as-is.
        name_dedup_map: Optional map from species_id to deduplicated name (with suffix if needed)
        
    Returns:
        Rule with IDs replaced by Names (quoted if needed) when use_name=True
    """
    if not use_name:
        return rule
    
    import re
    result = rule
    
    # Replace each species ID with its name
    for sid, sp in species.items():
        if not sp.name:
            continue
        
        # Use deduplicated name if map is provided
        if name_dedup_map and sid in name_dedup_map:
            name = name_dedup_map[sid]
        else:
            name = sp.name.strip()
        
        is_valid = is_valid_sid(name)
        replacement = name if is_valid else f'"{name}"'
        
        # Use word boundary pattern to match ID as whole token
        pattern = r'\b' + re.escape(sid) + r'(?=\s|&|\||!|\(|\)|>=|<=|>|<|!=|=|:|$)'
        result = re.sub(pattern, replacement, result)
    
    return result

def write_spreadsheet(model: QualModel, output_path: str, template_path: str = None, rule_format: str = "operators", use_name: bool = False) -> str:
    """Write QualModel to spreadsheet format
    
    Args:
        model: The in-memory model to write
        output_path: Path to output spreadsheet file
        template_path: Optional path to template.xlsx for README and Appendix sheets
        rule_format: Format for transition rules - "operators" (default, uses >=, <=, etc.) or "colon" (uses : notation)
    
    Returns:
        str: The actual output path used (may have .xlsx added)
    """
    # Ensure output path has .xlsx extension
    output_p = Path(output_path)
    if output_p.suffix.lower() not in ('.xlsx', '.xls', '.xlsm'):
        output_path = str(output_p) + '.xlsx'
    
    # Use template if provided, otherwise create new workbook
    if template_path and Path(template_path).exists():
        # Copy template to output
        shutil.copy(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        
        # Remove data sheets if they exist, keep README and Appendix
        for sheet_name in ['Model', 'Species', 'Transitions', 'Interactions']:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    
    # Determine insertion position (after README if it exists)
    insert_pos = 1 if 'README' in wb.sheetnames else 0
    
    # Build name deduplication map when use_name=True (used across all sheets)
    name_dedup_map = {}
    if use_name:
        name_dedup_map = _build_name_deduplication_map(model.species)
    
    # Create sheets in order: README, Model, Species, Interactions, Transitions, Appendix
    _write_model_sheet(wb, model, insert_pos)
    _write_species_sheet(wb, model, insert_pos + 1, use_name=use_name, name_dedup_map=name_dedup_map)
    _write_interactions_sheet(wb, model, insert_pos + 2, use_name=use_name, name_dedup_map=name_dedup_map)
    _write_transitions_sheet(wb, model, insert_pos + 3, rule_format, use_name=use_name, name_dedup_map=name_dedup_map)
    
    # Save
    wb.save(output_path)
    wb.close()
    
    return output_path


def _write_model_sheet(wb: openpyxl.Workbook, model: QualModel, position: int = 0):
    """Write Model sheet"""
    ws = wb.create_sheet("Model", position)
    
    # No headers for Model sheet
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Data rows (start from row 1)
    row = 1
    
    # Model_source
    ws.cell(row=row, column=1, value="Model_source")
    if model.model.source_urls:
        ws.cell(row=row, column=2, value=", ".join(_url_to_compact_id(url) for url in model.model.source_urls))
    row += 1
    
    # Model_ID (required)
    cell = ws.cell(row=row, column=1, value="Model_ID")
    cell.fill = required_fill
    ws.cell(row=row, column=2, value=model.model.model_id)
    row += 1
    
    # Name
    ws.cell(row=row, column=1, value="Name")
    if model.model.name:
        ws.cell(row=row, column=2, value=model.model.name)
    row += 1
    
    # Publication
    ws.cell(row=row, column=1, value="Publication")
    if model.model.described_by:
        ws.cell(row=row, column=2, value=", ".join(_url_to_compact_id(url) for url in model.model.described_by))
    row += 1
    
    # Origin_publication & Origin_model (from derived_from)
    origin_pubs = []
    origin_models = []
    if model.model.derived_from:
        for url in model.model.derived_from:
            compact_id = _url_to_compact_id(url)
            if 'biomodels' in compact_id.lower():
                origin_models.append(compact_id)
            else:
                origin_pubs.append(compact_id)
    
    ws.cell(row=row, column=1, value="Origin_publication")
    if origin_pubs:
        ws.cell(row=row, column=2, value=", ".join(origin_pubs))
    row += 1
    
    ws.cell(row=row, column=1, value="Origin_model")
    if origin_models:
        ws.cell(row=row, column=2, value=", ".join(origin_models))
    row += 1
    
    # Taxon
    ws.cell(row=row, column=1, value="Taxon")
    if model.model.taxons:
        ws.cell(row=row, column=2, value=", ".join(_url_to_compact_id(taxon) for taxon in model.model.taxons))
    row += 1
    
    # Biological_process
    ws.cell(row=row, column=1, value="Biological_process")
    if model.model.biological_processes:
        ws.cell(row=row, column=2, value=", ".join(_url_to_compact_id(process) for process in model.model.biological_processes))
    row += 1
    
    # Created
    ws.cell(row=row, column=1, value="Created")
    created_value = model.model.created_iso
    if not created_value:
        # Fallback to current time if not set
        from datetime import datetime, timezone
        created_value = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    ws.cell(row=row, column=2, value=created_value)
    row += 1
    
    # Modified - use current timestamp
    from datetime import datetime, timezone
    ws.cell(row=row, column=1, value="Modified")
    ws.cell(row=row, column=2, value=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"))
    row += 1
    
    # Creators
    max_creators = max(2, len(model.model.creators))
    for idx in range(1, max_creators + 1):
        ws.cell(row=row, column=1, value=f"Creator{idx}")
        if idx <= len(model.model.creators):
            ws.cell(row=row, column=2, value=_person_to_string(model.model.creators[idx - 1]))
        row += 1
    
    # Contributors
    max_contributors = max(2, len(model.model.contributors))
    for idx in range(1, max_contributors + 1):
        ws.cell(row=row, column=1, value=f"Contributor{idx}")
        if idx <= len(model.model.contributors):
            ws.cell(row=row, column=2, value=_person_to_string(model.model.contributors[idx - 1]))
        row += 1
    
    # Version
    ws.cell(row=row, column=1, value="Version")
    if model.model.versions:
        # Format as "Version: version1, version2, ..." if multiple versions
        version_str = ", ".join(model.model.versions)
        if len(model.model.versions) > 1 or not version_str.startswith("Version:"):
            version_str = f"Version: {version_str}"
        ws.cell(row=row, column=2, value=version_str)
    row += 1
    
    # Notes
    max_notes = max(3, len(model.model.notes))
    for idx in range(1, max_notes + 1):
        ws.cell(row=row, column=1, value=f"Notes{idx}")
        if idx <= len(model.model.notes):
            ws.cell(row=row, column=2, value=model.model.notes[idx - 1])
        row += 1
    
    # Comments - add TabularQual version
    ws.cell(row=row, column=1, value="Comments")
    ws.cell(row=row, column=2, value=f"Created by TabularQual version {spec.VERSION}")
    row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80


def _write_species_sheet(wb: openpyxl.Workbook, model: QualModel, position: int = 0, 
                         use_name: bool = False, name_dedup_map: Dict[str, str] = None):
    """Write Species sheet
    
    Args:
        use_name: If True, deduplicate names by adding suffixes when writing
        name_dedup_map: Map from species_id to deduplicated name (with suffix if needed)
    """
    ws = wb.create_sheet("Species", position)
    
    # Header style
    header_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    header_font = Font(bold=True)
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Group annotations by qualifier and determine max unique qualifiers
    max_qualifiers = 0
    max_notes = 0
    for species in model.species.values():
        # Group by qualifier
        grouped = {}
        for qualifier, identifier in species.annotations:
            if qualifier not in grouped:
                grouped[qualifier] = []
            grouped[qualifier].append(identifier)
        max_qualifiers = max(max_qualifiers, len(grouped))
        max_notes = max(max_notes, len(species.notes))
    
    # Ensure at least 2 Relation/Identifier pairs
    max_qualifiers = max(max_qualifiers, 2)
    # Ensure at least Notes1 column
    max_notes = max(max_notes, 1)
    
    # Build headers
    headers = ["Species_ID", "Name"]
    
    # Add Relation/Identifier pairs
    for i in range(max_qualifiers):
        headers.extend([f"Relation{i + 1}", f"Identifier{i + 1}"])
    
    headers.extend(["Compartment", "Type", "Constant", "InitialLevel", "MaxLevel"])
    
    # Add Notes
    for i in range(max_notes):
        headers.append(f"Notes{i + 1}")
    
    headers.append("Comments")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        if header == "Species_ID":
            cell.fill = required_fill
    
    # Write species data
    for row_idx, species in enumerate(sorted(model.species.values(), key=lambda s: s.species_id), 2):
        col = 1
        
        # Species_ID
        ws.cell(row=row_idx, column=col, value=species.species_id)
        col += 1
        
        # Name - use deduplicated version if use_name=True
        if use_name and species.name:
            name_value = name_dedup_map.get(species.species_id, species.name.strip())
        else:
            name_value = species.name.strip() if species.name else None
        ws.cell(row=row_idx, column=col, value=name_value)
        col += 1
        
        # Annotations - group by qualifier and combine identifiers
        grouped_annos = {}
        for qualifier, identifier in species.annotations:
            if qualifier not in grouped_annos:
                grouped_annos[qualifier] = []
            grouped_annos[qualifier].append(_url_to_compact_id(identifier))
        
        for i in range(max_qualifiers):
            if i < len(grouped_annos):
                qualifier = list(grouped_annos.keys())[i]
                identifiers = grouped_annos[qualifier]
                ws.cell(row=row_idx, column=col, value=qualifier)
                ws.cell(row=row_idx, column=col + 1, value=", ".join(identifiers))
            col += 2
        
        # Compartment
        ws.cell(row=row_idx, column=col, value=species.compartment)
        col += 1
        
        # Type
        ws.cell(row=row_idx, column=col, value=species.type)
        col += 1
        
        # Constant
        if species.constant is not None:
            ws.cell(row=row_idx, column=col, value=str(species.constant))
        col += 1
        
        # InitialLevel
        if species.initial_level is not None:
            ws.cell(row=row_idx, column=col, value=species.initial_level)
        col += 1
        
        # MaxLevel
        if species.max_level is not None:
            ws.cell(row=row_idx, column=col, value=species.max_level)
        col += 1
        
        # Notes - write each note to separate column
        for i in range(max_notes):
            if i < len(species.notes):
                ws.cell(row=row_idx, column=col, value=species.notes[i])
            col += 1
        
        # Comments (skip)
        col += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30


def _write_transitions_sheet(wb: openpyxl.Workbook, model: QualModel, position: int = 0, 
                            rule_format: str = "operators", use_name: bool = False, 
                            name_dedup_map: Dict[str, str] = None):
    """Write Transitions sheet"""
    ws = wb.create_sheet("Transitions", position)
    
    # Header style
    header_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    header_font = Font(bold=True)
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Group annotations by qualifier and determine max unique qualifiers
    max_qualifiers = 0
    max_notes = 0
    for transition in model.transitions:
        # Group by qualifier
        grouped = {}
        for qualifier, identifier in transition.annotations:
            if qualifier not in grouped:
                grouped[qualifier] = []
            grouped[qualifier].append(identifier)
        max_qualifiers = max(max_qualifiers, len(grouped))
        max_notes = max(max_notes, len(transition.notes))
    
    # Ensure at least 1 Relation/Identifier pair
    max_qualifiers = max(max_qualifiers, 1)
    # Ensure at least Notes1 column
    max_notes = max(max_notes, 1)
    
    # Build headers
    headers = ["Transitions_ID", "Name", "Target", "Level", "Rule"]
    
    # Add Relation/Identifier pairs
    for i in range(max_qualifiers):
        headers.extend([f"Relation{i + 1}", f"Identifier{i + 1}"])
    
    # Add Notes
    for i in range(max_notes):
        headers.append(f"Notes{i + 1}")
    
    headers.append("Comments")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        if header in ["Target", "Rule"]:
            cell.fill = required_fill
    
    # Write transition data
    for row_idx, transition in enumerate(model.transitions, 2):
        col = 1
        
        # Transitions_ID
        ws.cell(row=row_idx, column=col, value=transition.transition_id)
        col += 1
        
        # Name
        ws.cell(row=row_idx, column=col, value=transition.name)
        col += 1
        
        # Target - use Name if use_name=False
        target_value = _resolve_id_to_name_for_output(transition.target, model.species, use_name, name_dedup_map)
        ws.cell(row=row_idx, column=col, value=target_value)
        col += 1
        
        # Level
        if transition.level is not None:
            ws.cell(row=row_idx, column=col, value=transition.level)
        col += 1
        
        # Rule - convert format if needed, and resolve IDs to Names if use_name=False
        rule = transition.rule
        if use_name:
            rule = _resolve_rule_ids_to_names(rule, model.species, use_name, name_dedup_map)
        if rule_format == "colon":
            rule = _convert_rule_to_colon(rule)
        ws.cell(row=row_idx, column=col, value=rule)
        col += 1
        
        # Annotations - group by qualifier and combine identifiers
        grouped_annos = {}
        for qualifier, identifier in transition.annotations:
            if qualifier not in grouped_annos:
                grouped_annos[qualifier] = []
            grouped_annos[qualifier].append(_url_to_compact_id(identifier))
        
        for i in range(max_qualifiers):
            if i < len(grouped_annos):
                qualifier = list(grouped_annos.keys())[i]
                identifiers = grouped_annos[qualifier]
                ws.cell(row=row_idx, column=col, value=qualifier)
                ws.cell(row=row_idx, column=col + 1, value=", ".join(identifiers))
            col += 2
        
        # Notes
        for i in range(max_notes):
            if i < len(transition.notes):
                ws.cell(row=row_idx, column=col, value=transition.notes[i])
            col += 1
        
        # Comments
        col += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['E'].width = 40


def _write_interactions_sheet(wb: openpyxl.Workbook, model: QualModel, position: int = 0, 
                              use_name: bool = False, name_dedup_map: Dict[str, str] = None):
    """Write Interactions sheet"""
    # Always create the sheet, even if no interactions
    ws = wb.create_sheet("Interactions", position)
    
    # Header style
    header_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    header_font = Font(bold=True)
    
    # Group annotations by qualifier and determine max unique qualifiers
    max_qualifiers = 0
    max_notes = 0
    for interaction in model.interactions:
        # Group by qualifier
        grouped = {}
        for qualifier, identifier in interaction.annotations:
            if qualifier not in grouped:
                grouped[qualifier] = []
            grouped[qualifier].append(identifier)
        max_qualifiers = max(max_qualifiers, len(grouped))
        max_notes = max(max_notes, len(interaction.notes))
    
    # Ensure at least 1 Relation/Identifier pair
    max_qualifiers = max(max_qualifiers, 1)
    # Ensure at least Notes1 column
    max_notes = max(max_notes, 1)
    
    # Build headers
    headers = ["Target", "Source", "Sign"]
    
    # Add Relation/Identifier pairs
    for i in range(max_qualifiers):
        headers.extend([f"Relation{i + 1}", f"Identifier{i + 1}"])
    
    # Add Notes
    for i in range(max_notes):
        headers.append(f"Notes{i + 1}")
    
    headers.append("Comments")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Write interaction data
    for row_idx, interaction in enumerate(model.interactions, 2):
        col = 1
        
        # Target - use Name if use_name=False
        target_value = _resolve_id_to_name_for_output(interaction.target, model.species, use_name, name_dedup_map)
        ws.cell(row=row_idx, column=col, value=target_value)
        col += 1
        
        # Source - use Name if use_name=False
        source_value = _resolve_id_to_name_for_output(interaction.source, model.species, use_name, name_dedup_map)
        ws.cell(row=row_idx, column=col, value=source_value)
        col += 1
        
        # Sign
        ws.cell(row=row_idx, column=col, value=interaction.sign)
        col += 1
        
        # Annotations - group by qualifier and combine identifiers
        grouped_annos = {}
        for qualifier, identifier in interaction.annotations:
            if qualifier not in grouped_annos:
                grouped_annos[qualifier] = []
            grouped_annos[qualifier].append(_url_to_compact_id(identifier))
        
        for i in range(max_qualifiers):
            if i < len(grouped_annos):
                qualifier = list(grouped_annos.keys())[i]
                identifiers = grouped_annos[qualifier]
                ws.cell(row=row_idx, column=col, value=qualifier)
                ws.cell(row=row_idx, column=col + 1, value=", ".join(identifiers))
            col += 2
        
        # Notes
        for i in range(max_notes):
            if i < len(interaction.notes):
                ws.cell(row=row_idx, column=col, value=interaction.notes[i])
            col += 1
        
        # Comments
        col += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15


def _url_to_compact_id(url: str) -> str:
    """Convert identifiers.org URL or URN to compact ID
    
    Handles:
    - identifiers.org/ncbigene:7132 -> ncbigene:7132
    - identifiers.org/ncbigene/7132 -> ncbigene:7132 (slash to colon)
    - urn:miriam:ncbigene:596 -> ncbigene:596
    - Other URLs are kept as-is
    """
    # Handle identifiers.org URLs
    if "identifiers.org/" in url:
        # Extract the part after identifiers.org/
        parts = url.split("identifiers.org/")
        if len(parts) > 1:
            compact_id = parts[1]
            # Convert slash to colon if present (e.g., ncbigene/7132 -> ncbigene:7132)
            # Only convert the first slash to maintain compatibility with nested identifiers
            if "/" in compact_id and ":" not in compact_id:
                compact_id = compact_id.replace("/", ":", 1)
            return compact_id
    
    # Handle urn:miriam: format (e.g., urn:miriam:ncbigene:596 -> ncbigene:596)
    if url.startswith("urn:miriam:"):
        return url[len("urn:miriam:"):]
    
    return url


def _person_to_string(person: Person) -> str:
    """Convert Person object to formatted string"""
    parts = []
    if person.family_name:
        parts.append(person.family_name)
    if person.given_name:
        parts.append(person.given_name)
    if person.organization:
        parts.append(f'"{person.organization}"')
    if person.email:
        parts.append(person.email)
    return ", ".join(parts) if parts else ""


def _convert_rule_to_colon(rule: str) -> str:
    """Convert rule from operator format to colon format
    
    Examples:
        A >= 2 -> A:2
        B < 2 -> !B:2
        C = 0 -> !C
        D = 1 -> D
        E >= 1 -> E
    """
    import re
    
    # Handle >= operator (replace with colon notation)
    # A >= 2 becomes A:2, but A >= 1 becomes A
    def replace_geq(match):
        var = match.group(1)
        threshold = match.group(2)
        if threshold == '1':
            return var
        return f"{var}:{threshold}"
    
    rule = re.sub(r'(\w+)\s*>=\s*(\d+)', replace_geq, rule)
    
    # Handle < operator (replace with negation and colon notation)
    # A < 2 becomes !A:2, but A < 1 becomes !A
    def replace_lt(match):
        var = match.group(1)
        threshold = match.group(2)
        if threshold == '1':
            return f"!{var}"
        return f"!{var}:{threshold}"
    
    rule = re.sub(r'(\w+)\s*<\s*(\d+)', replace_lt, rule)
    
    # Handle = operator
    # A = 0 becomes !A, A = 1 becomes A, A = N becomes A:N
    def replace_eq(match):
        var = match.group(1)
        threshold = match.group(2)
        if threshold == '0':
            return f"!{var}"
        elif threshold == '1':
            return var
        return f"{var}:{threshold}"
    
    rule = re.sub(r'(\w+)\s*=\s*(\d+)', replace_eq, rule)
    
    # Handle == operator (same as =)
    def replace_eqeq(match):
        var = match.group(1)
        threshold = match.group(2)
        if threshold == '0':
            return f"!{var}"
        elif threshold == '1':
            return var
        return f"{var}:{threshold}"
    
    rule = re.sub(r'(\w+)\s*==\s*(\d+)', replace_eqeq, rule)
    
    return rule


def write_csv(model: QualModel, output_prefix: str, rule_format: str = "operators", use_name: bool = False):
    # Build name deduplication map when use_name=True (used across all CSV files)
    name_dedup_map = {}
    if use_name:
        name_dedup_map = _build_name_deduplication_map(model.species)
    """Write QualModel to CSV files.
    
    Args:
        model: The in-memory model to write
        output_prefix: Prefix for output CSV files (e.g., 'Example' creates Example_Model.csv, etc.)
        rule_format: Format for transition rules - "operators" or "colon"
    
    Returns:
        List of created file paths
    """
    output_path = Path(output_prefix)
    output_dir = output_path.parent if output_path.parent.exists() else Path('.')
    prefix = output_path.name
    
    created_files = []
    
    # Write Model CSV
    model_file = output_dir / f"{prefix}_Model.csv"
    _write_model_csv(model, str(model_file))
    created_files.append(str(model_file))
    
    # Write Species CSV
    species_file = output_dir / f"{prefix}_Species.csv"
    _write_species_csv(model, str(species_file), use_name=use_name, name_dedup_map=name_dedup_map)
    created_files.append(str(species_file))
    
    # Write Transitions CSV
    transitions_file = output_dir / f"{prefix}_Transitions.csv"
    _write_transitions_csv(model, str(transitions_file), rule_format, use_name=use_name)
    created_files.append(str(transitions_file))
    
    # Write Interactions CSV
    interactions_file = output_dir / f"{prefix}_Interactions.csv"
    _write_interactions_csv(model, str(interactions_file), use_name=use_name)
    created_files.append(str(interactions_file))
    
    return created_files


def _write_model_csv(model: QualModel, output_path: str):
    """Write Model sheet to CSV (vertical format: key-value pairs)"""
    from datetime import datetime, timezone
    
    rows = []
    
    # Model_source
    source_val = ", ".join(_url_to_compact_id(url) for url in model.model.source_urls) if model.model.source_urls else ""
    rows.append(["Model_source", source_val])
    
    # Model_ID (required)
    rows.append(["Model_ID", model.model.model_id])
    
    # Name
    rows.append(["Name", model.model.name or ""])
    
    # Publication
    pub_val = ", ".join(_url_to_compact_id(url) for url in model.model.described_by) if model.model.described_by else ""
    rows.append(["Publication", pub_val])
    
    # Origin_publication & Origin_model (from derived_from)
    origin_pubs = []
    origin_models = []
    if model.model.derived_from:
        for url in model.model.derived_from:
            compact_id = _url_to_compact_id(url)
            if 'biomodels' in compact_id.lower():
                origin_models.append(compact_id)
            else:
                origin_pubs.append(compact_id)
    
    rows.append(["Origin_publication", ", ".join(origin_pubs)])
    rows.append(["Origin_model", ", ".join(origin_models)])
    
    # Taxon
    taxon_val = ", ".join(_url_to_compact_id(taxon) for taxon in model.model.taxons) if model.model.taxons else ""
    rows.append(["Taxon", taxon_val])
    
    # Biological_process
    bp_val = ", ".join(_url_to_compact_id(process) for process in model.model.biological_processes) if model.model.biological_processes else ""
    rows.append(["Biological_process", bp_val])
    
    # Created
    created_value = model.model.created_iso or datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    rows.append(["Created", created_value])
    
    # Modified
    rows.append(["Modified", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")])
    
    # Creators
    max_creators = max(2, len(model.model.creators))
    for idx in range(1, max_creators + 1):
        value = _person_to_string(model.model.creators[idx - 1]) if idx <= len(model.model.creators) else ""
        rows.append([f"Creator{idx}", value])
    
    # Contributors
    max_contributors = max(2, len(model.model.contributors))
    for idx in range(1, max_contributors + 1):
        value = _person_to_string(model.model.contributors[idx - 1]) if idx <= len(model.model.contributors) else ""
        rows.append([f"Contributor{idx}", value])
    
    # Version
    version_str = ""
    if model.model.versions:
        version_str = ", ".join(model.model.versions)
        if len(model.model.versions) > 1 or not version_str.startswith("Version:"):
            version_str = f"Version: {version_str}"
    rows.append(["Version", version_str])
    
    # Notes
    max_notes = max(3, len(model.model.notes))
    for idx in range(1, max_notes + 1):
        value = model.model.notes[idx - 1] if idx <= len(model.model.notes) else ""
        rows.append([f"Notes{idx}", value])
    
    # Comments
    rows.append(["Comments", f"Created by TabularQual version {spec.VERSION}"])
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def _write_species_csv(model: QualModel, output_path: str, use_name: bool = False, name_dedup_map: Dict[str, str] = None):
    """Write Species sheet to CSV"""
    # Determine max qualifiers and notes
    max_qualifiers = 0
    max_notes = 0
    for species in model.species.values():
        grouped = {}
        for qualifier, identifier in species.annotations:
            if qualifier not in grouped:
                grouped[qualifier] = []
            grouped[qualifier].append(identifier)
        max_qualifiers = max(max_qualifiers, len(grouped))
        max_notes = max(max_notes, len(species.notes))
    
    max_qualifiers = max(max_qualifiers, 2)
    max_notes = max(max_notes, 1)
    
    # Build headers
    headers = ["Species_ID", "Name"]
    for i in range(max_qualifiers):
        headers.extend([f"Relation{i + 1}", f"Identifier{i + 1}"])
    headers.extend(["Compartment", "Type", "Constant", "InitialLevel", "MaxLevel"])
    for i in range(max_notes):
        headers.append(f"Notes{i + 1}")
    headers.append("Comments")
    
    rows = [headers]
    
    for species in sorted(model.species.values(), key=lambda s: s.species_id):
        # Name - use deduplicated version if use_name=True
        if use_name and species.name:
            name_value = name_dedup_map.get(species.species_id, species.name.strip()) if name_dedup_map else species.name.strip()
        else:
            name_value = species.name.strip() if species.name else ""
        row = [species.species_id, name_value]
        
        # Annotations
        grouped_annos = {}
        for qualifier, identifier in species.annotations:
            if qualifier not in grouped_annos:
                grouped_annos[qualifier] = []
            grouped_annos[qualifier].append(_url_to_compact_id(identifier))
        
        for i in range(max_qualifiers):
            if i < len(grouped_annos):
                qualifier = list(grouped_annos.keys())[i]
                identifiers = grouped_annos[qualifier]
                row.extend([qualifier, ", ".join(identifiers)])
            else:
                row.extend(["", ""])
        
        row.extend([
            species.compartment or "",
            species.type or "",
            str(species.constant) if species.constant is not None else "",
            species.initial_level if species.initial_level is not None else "",
            species.max_level if species.max_level is not None else ""
        ])
        
        for i in range(max_notes):
            row.append(species.notes[i] if i < len(species.notes) else "")
        
        row.append("")  # Comments
        rows.append(row)
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def _write_transitions_csv(model: QualModel, output_path: str, rule_format: str = "operators", 
                          use_name: bool = False, name_dedup_map: Dict[str, str] = None):
    """Write Transitions sheet to CSV"""
    # Determine max qualifiers and notes
    max_qualifiers = 0
    max_notes = 0
    for transition in model.transitions:
        grouped = {}
        for qualifier, identifier in transition.annotations:
            if qualifier not in grouped:
                grouped[qualifier] = []
            grouped[qualifier].append(identifier)
        max_qualifiers = max(max_qualifiers, len(grouped))
        max_notes = max(max_notes, len(transition.notes))
    
    max_qualifiers = max(max_qualifiers, 1)
    max_notes = max(max_notes, 1)
    
    # Build headers
    headers = ["Transitions_ID", "Name", "Target", "Level", "Rule"]
    for i in range(max_qualifiers):
        headers.extend([f"Relation{i + 1}", f"Identifier{i + 1}"])
    for i in range(max_notes):
        headers.append(f"Notes{i + 1}")
    headers.append("Comments")
    
    rows = [headers]
    
    for transition in model.transitions:
        # Target - use Name if use_name=False
        target_value = _resolve_id_to_name_for_output(transition.target, model.species, use_name, name_dedup_map)
        
        # Rule - convert format if needed, and resolve IDs to Names if use_name=False
        rule = transition.rule
        if use_name:
            rule = _resolve_rule_ids_to_names(rule, model.species, use_name, name_dedup_map)
        if rule_format == "colon":
            rule = _convert_rule_to_colon(rule)
        
        row = [
            transition.transition_id or "",
            transition.name or "",
            target_value,
            transition.level if transition.level is not None else "",
            rule
        ]
        
        # Annotations
        grouped_annos = {}
        for qualifier, identifier in transition.annotations:
            if qualifier not in grouped_annos:
                grouped_annos[qualifier] = []
            grouped_annos[qualifier].append(_url_to_compact_id(identifier))
        
        for i in range(max_qualifiers):
            if i < len(grouped_annos):
                qualifier = list(grouped_annos.keys())[i]
                identifiers = grouped_annos[qualifier]
                row.extend([qualifier, ", ".join(identifiers)])
            else:
                row.extend(["", ""])
        
        for i in range(max_notes):
            row.append(transition.notes[i] if i < len(transition.notes) else "")
        
        row.append("")  # Comments
        rows.append(row)
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def _write_interactions_csv(model: QualModel, output_path: str, use_name: bool = False, 
                           name_dedup_map: Dict[str, str] = None):
    """Write Interactions sheet to CSV"""
    # Determine max qualifiers and notes
    max_qualifiers = 0
    max_notes = 0
    for interaction in model.interactions:
        grouped = {}
        for qualifier, identifier in interaction.annotations:
            if qualifier not in grouped:
                grouped[qualifier] = []
            grouped[qualifier].append(identifier)
        max_qualifiers = max(max_qualifiers, len(grouped))
        max_notes = max(max_notes, len(interaction.notes))
    
    max_qualifiers = max(max_qualifiers, 1)
    max_notes = max(max_notes, 1)
    
    # Build headers
    headers = ["Target", "Source", "Sign"]
    for i in range(max_qualifiers):
        headers.extend([f"Relation{i + 1}", f"Identifier{i + 1}"])
    for i in range(max_notes):
        headers.append(f"Notes{i + 1}")
    headers.append("Comments")
    
    rows = [headers]
    
    for interaction in model.interactions:
        # Resolve target and source to Names if use_name=False
        target_value = _resolve_id_to_name_for_output(interaction.target, model.species, use_name, name_dedup_map)
        source_value = _resolve_id_to_name_for_output(interaction.source, model.species, use_name, name_dedup_map)
        
        row = [
            target_value,
            source_value,
            interaction.sign or ""
        ]
        
        # Annotations
        grouped_annos = {}
        for qualifier, identifier in interaction.annotations:
            if qualifier not in grouped_annos:
                grouped_annos[qualifier] = []
            grouped_annos[qualifier].append(_url_to_compact_id(identifier))
        
        for i in range(max_qualifiers):
            if i < len(grouped_annos):
                qualifier = list(grouped_annos.keys())[i]
                identifiers = grouped_annos[qualifier]
                row.extend([qualifier, ", ".join(identifiers)])
            else:
                row.extend(["", ""])
        
        for i in range(max_notes):
            row.append(interaction.notes[i] if i < len(interaction.notes) else "")
        
        row.append("")  # Comments
        rows.append(row)
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(rows)