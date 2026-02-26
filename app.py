"""
TabularQual Web converter
A Streamlit web application for converting between SpreadSBML spreadsheets and SBML-qual files.
"""

import streamlit as st
import tempfile
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import warnings
from io import BytesIO
import gc

# Import tabularqual functions
from tabularqual.convert_spreadsheet_to_sbml import convert_spreadsheet_to_sbml
from tabularqual.convert_sbml_to_spreadsheet import convert_sbml_to_spreadsheet, get_default_template_path
from tabularqual import spec
import zipfile

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Page configuration
st.set_page_config(
    page_title="TabularQual",
    page_icon="🔄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
    <style>
    .main-header {
        font-size: 3.5rem;
        font-weight: bold;
        color: #1f77b4;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        /* font-size: 2.5rem; */
        color: #666;
        margin-bottom: 2rem;
    }
    h2 {
        font-size: 1rem;
    }
    .stDownloadButton button {
        background-color: #28a745;
        color: white;
        font-weight: bold;
    }
    .success-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        margin: 1rem 0;
    }
    .info-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        color: #0c5460;
        margin: 1rem 0;
    }
    </style>
""", unsafe_allow_html=True)

# Header
st.markdown('<p class="main-header">🔄 TabularQual</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Convert between spreadsheets and SBML-qual files for logical models (Boolean and multi-valued). <br>For more about the format, see the <a href="https://docs.google.com/document/d/1RCIN4bOsw4Uq9X2I-gdfBXDydyViYzaVhQK8cpdEWhA/edit?usp=sharing">Spreadsheet specification</a></p>', unsafe_allow_html=True)

# Sidebar with information
with st.sidebar:
    st.header("ℹ️ About")
    st.markdown("""
    This tool converts between:
    - **Spreadsheets** (.xlsx or .csv)
    - **SBML-qual** files (.sbml/.xml)
    """)
    st.header("🔍 Examples")
    st.markdown("""
    - [Faure2006 (Boolean)](https://docs.google.com/spreadsheets/d/1B9SUcuY_ioQVlY9y351yIHnW45oZ8J1t/edit?usp=drive_link&ouid=105819375684543832411&rtpof=true&sd=true)
    - [ThieffryThomas1995 (Multi-valued)](https://docs.google.com/spreadsheets/d/1Auepvb1Z0Q4lIjMqaesjWdh3oHWTwjA8/edit?usp=drive_link&ouid=105819375684543832411&rtpof=true&sd=true)
    """)
    
    st.header("📖 Documentation")
    st.markdown("""
    **Transition Rules Syntax:**
    - Logical: `&` (AND), `|` (OR), `!` (NOT), `^` (XOR)
    - Comparisons: `>=`, `<=`, `<`, `>`, `!=`
    - Colon symbol: `A:2` means `A >= 2`
    - Negated: `!A:2` means `A < 2`
    - Constant rules: `FALSE` / `TRUE`/ `N` (integer) means the target will be fixed at level 0 / 1 / N

    """)
    
    st.header("🔗 Links")
    st.markdown("""
    - [GitHub Repository](https://github.com/sys-bio/TabularQual)
    - [The Spreadsheet Specification](https://docs.google.com/document/d/1RCIN4bOsw4Uq9X2I-gdfBXDydyViYzaVhQK8cpdEWhA/edit?usp=sharing)
    - [Relevant files](https://drive.google.com/drive/folders/14lE0jmL4wPnwbfdwgPTs22URD32sovjw?usp=drive_link)
    """)

# Main content
tab1, tab2 = st.tabs(["📊 Spreadsheet → SBML", "📋 SBML → Spreadsheet"])

# Tab 1: Spreadsheet to SBML
with tab1:
    # st.header("Convert Spreadsheet to SBML")
    st.markdown('<h2 style="font-size:27px;">Convert Spreadsheet to SBML</h2>', unsafe_allow_html=True)
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        uploaded_files = st.file_uploader(
            "Upload Excel (.xlsx) or CSV files. For CSV, upload Species.csv and Transitions.csv (required), plus Model.csv and Interactions.csv (optional). You may use this [Template](https://docs.google.com/spreadsheets/d/1_welMPd8-Wdbu3fTrCjUZz159yT5kxNO/edit?usp=sharing&ouid=105819375684543832411&rtpof=true&sd=true).",
            type=["xlsx", "csv"],
            key="spreadsheet_upload",
            accept_multiple_files=True,
            help="Upload a single XLSX file or multiple CSV files"
        )
        
        # Auto-detect format and separate files
        uploaded_xlsx = None
        uploaded_csvs = []
        if uploaded_files:
            for f in uploaded_files:
                if f.name.lower().endswith('.xlsx'):
                    uploaded_xlsx = f
                elif f.name.lower().endswith('.csv'):
                    uploaded_csvs.append(f)
            
            # Show what was detected
            if uploaded_xlsx:
                st.caption(f"📗 Excel file: {uploaded_xlsx.name}")
            if uploaded_csvs:
                csv_names = [f.name for f in uploaded_csvs]
                st.caption(f"📄 CSV files: {', '.join(csv_names)}")
    
    with col2:
        st.markdown("### Options")
        inter_anno = st.checkbox(
            "Include Interaction Annotations",
            value=True,
            key="inter_anno",
            help="Include annotations from the Interaction tab in the SBML output"
        )
        trans_anno = st.checkbox(
            "Include Transition Annotations",
            value=True,
            key="trans_anno",
            help="Include annotations from the Transitions tab in the SBML output"
        )
        validate_annotations = st.checkbox(
            "Validate Annotations",
            value=True,
            key="validate_anno_tab1",
            help="Validate SBML annotations using sbmlutils (requires sbmlutils with metadata.validator)"
        )
        use_name = st.checkbox(
            "Use Species Name",
            value=False,
            key="use_name_tab1",
            help="Species Name has been used in rules and interactions. If unchecked (default), Species_ID."
        )
    
    # Determine if we have valid input (prefer XLSX if both provided)
    has_valid_input = uploaded_xlsx is not None or len(uploaded_csvs) > 0
    
    if has_valid_input:
        # Preview and handle input based on format (prefer XLSX if both provided)
        if uploaded_xlsx is not None:
            # Store file content once
            file_content = uploaded_xlsx.getvalue()
            original_name = uploaded_xlsx.name.rsplit('.', 1)[0]
            
            # Preview uploaded spreadsheet
            with st.expander("📊 Preview Uploaded Spreadsheet", expanded=False):
                try:
                    # Use read_only and data_only for memory efficiency
                    wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
                    
                    # Skip README and Appendix sheets
                    skip_sheets = ["README", "Appendix"]
                    skipped_sheets = [s for s in wb.sheetnames if s in skip_sheets]
                    
                    if skipped_sheets:
                        st.info(f"ℹ️ Skipping sheets: {', '.join(skipped_sheets)}")
                    
                    for sheet_name in wb.sheetnames:
                        if sheet_name in skip_sheets:
                            continue
                        
                        st.subheader(f"Sheet: {sheet_name}")
                        
                        # Read sheet data
                        sheet = wb[sheet_name]
                        data = []
                        max_preview_rows = 50
                        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                            if idx >= max_preview_rows:
                                break
                            data.append(row)
                        
                        if data:
                            # Find the maximum number of columns
                            max_cols = max(len(row) for row in data)
                            # Pad rows with None to make them equal length
                            # Convert all values to strings to avoid pyarrow serialization errors
                            padded_data = [
                                [str(v) if v is not None else "" for v in list(row) + [None] * (max_cols - len(row))]
                                for row in data
                            ]
                            df = pd.DataFrame(padded_data)
                            st.dataframe(df, width='stretch')
                            total_rows = sheet.max_row or 0
                            if idx >= max_preview_rows - 1 and total_rows > max_preview_rows:
                                st.info(f"Preview limited to {max_preview_rows} rows (sheet has {total_rows} rows)")
                            del df, padded_data, data
                            gc.collect()
                        else:
                            st.info("Sheet is empty")
                    
                    wb.close()
                    del wb  # Explicit cleanup
                    gc.collect()  # Force garbage collection
                except Exception as e:
                    st.error(f"Error previewing spreadsheet: {str(e)}")
        else:
            # CSV input - preview each file
            original_name = "model"  # Default name for CSV
            # Try to extract prefix from filenames
            for csv_file in uploaded_csvs:
                name = csv_file.name.rsplit('.', 1)[0]
                if '_' in name:
                    original_name = name.rsplit('_', 1)[0]
                    break
            
            with st.expander("📊 Preview Uploaded CSV Files", expanded=False):
                for csv_file in uploaded_csvs:
                    st.subheader(f"File: {csv_file.name}")
                    try:
                        df = pd.read_csv(csv_file, nrows=50)
                        st.dataframe(df, width='stretch')
                        csv_file.seek(0)  # Reset file pointer
                    except Exception as e:
                        st.error(f"Error previewing {csv_file.name}: {str(e)}")
        
        # Output filename editor
        st.markdown("### Output File Name")
        output_filename = st.text_input(
            "Edit output filename (without extension)",
            value=f"{original_name}_out",
            key="sbml_output_name",
            help="The file will be saved with .sbml extension"
        )
        
        # Convert button
        if st.button("🔄 Convert to SBML", type="primary", key="convert_to_sbml"):
            with st.spinner("Converting..."):
                try:
                    temp_files_to_cleanup = []
                    
                    if uploaded_xlsx is not None:
                        # XLSX input
                        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_in:
                            tmp_in.write(file_content)
                            input_path = tmp_in.name
                        temp_files_to_cleanup.append(input_path)
                    else:
                        # CSV input - save files to temp directory
                        temp_dir = tempfile.mkdtemp()
                        for csv_file in uploaded_csvs:
                            file_path = os.path.join(temp_dir, csv_file.name)
                            with open(file_path, 'wb') as f:
                                f.write(csv_file.getvalue())
                        input_path = temp_dir
                    
                    with tempfile.NamedTemporaryFile(suffix='.sbml', delete=False) as tmp_out:
                        output_path = tmp_out.name
                    temp_files_to_cleanup.append(output_path)
                    
                    stats = convert_spreadsheet_to_sbml(
                        input_path,
                        output_path,
                        interactions_anno=inter_anno,
                        transitions_anno=trans_anno,
                        print_messages=False,  # Display in app instead
                        validate=validate_annotations,
                        use_name=use_name
                    )
                    
                    # Cleanup CSV temp dir if created
                    if uploaded_xlsx is None:
                        import shutil
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    
                    species_count = stats['species']
                    transitions_count = stats['transitions']
                    interactions_count = stats['interactions']
                    all_messages = stats.get('warnings', [])
                    validation_errors = stats.get('validation_errors', [])
                    total_val_errors = stats.get('total_validation_errors', 0)

                    # Read output file
                    with open(output_path, 'r', encoding='utf-8') as f:
                        sbml_content = f.read()
                    
                    # Success message
                    st.markdown('<div class="success-box">✅ Conversion successful!</div>', unsafe_allow_html=True)
                    
                    # Display validation errors
                    if validation_errors:
                        with st.expander(f"❗️ Annotation Validation Warnings ({total_val_errors} total)", expanded=True):
                            for error in validation_errors:
                                st.warning(error)

                    # Display messages
                    if all_messages:
                        # Separate info messages from warnings and SId messages
                        info_msgs = [m for m in all_messages if m.startswith("Found ") or m.startswith("No ")]
                        sid_msgs = [m for m in all_messages if "Invalid SId" in m or "Duplicate ID" in m or "Invalid reference" in m]
                        use_name_msgs = [m for m in all_messages if "--use-name" in m or "Using ID mode" in m or "Using Name mode" in m]
                        warning_msgs = [m for m in all_messages if m not in info_msgs and m not in sid_msgs and m not in use_name_msgs]
                        
                        # Display SId validation messages in a dedicated expander
                        if sid_msgs:
                            with st.expander(f"🔧 SId Format Corrections ({len(sid_msgs)} total)", expanded=True):
                                for msg in sid_msgs:
                                    st.info(msg)
                        
                        # Display --use-name related messages in an expander
                        if use_name_msgs:
                            with st.expander(f"ℹ️ --use-name Flag Messages ({len(use_name_msgs)} total)", expanded=False):
                                for msg in use_name_msgs:
                                    st.info(msg)
                                                
                        if warning_msgs:
                            with st.expander(f"⚠️ Warnings ({len(warning_msgs)} total)", expanded=True):
                                for msg in warning_msgs:
                                    st.info(f"{msg}")
                    
                    # Display statistics
                    col1, col2, col3 = st.columns(3)
                    col1.metric("Species", species_count)
                    col2.metric("Transitions", transitions_count)
                    col3.metric("Interactions", interactions_count)
                    
                    # LIMIT preview size
                    with st.expander("📄 Preview SBML Output", expanded=False):
                        if len(sbml_content) > 50000:  # ~50KB
                            preview_lines = sbml_content.split('\n')[:100]
                            st.code('\n'.join(preview_lines) + "\n\n... (truncated, download to see full file)", language="xml")
                        else:
                            st.code(sbml_content, language="xml")
                    
                    # Download button
                    final_filename = f"{output_filename}.sbml" if output_filename else f"{original_name}_out.sbml"
                    st.download_button(
                        label="⬇️ Download SBML File",
                        data=sbml_content,
                        file_name=final_filename,
                        mime="application/xml",
                        type="primary"
                    )
                    
                    # Cleanup
                    del sbml_content
                    for f in temp_files_to_cleanup:
                        try:
                            os.unlink(f)
                        except:
                            pass
                    gc.collect()
                    
                except Exception as e:
                    st.error(f"❌ Conversion failed: {str(e)}")
                    import traceback
                    with st.expander("Error Details"):
                        st.code(traceback.format_exc())
                    gc.collect()

# Tab 2: SBML to Spreadsheet
with tab2:
    # st.header("Convert SBML to Spreadsheet")
    st.markdown('<h2 style="font-size:27px;">Convert SBML to Spreadsheet</h2>', unsafe_allow_html=True)
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        uploaded_sbml = st.file_uploader(
            "Upload SBML-qual file (.sbml or .xml)",
            type=["sbml", "xml"],
            key="sbml_upload",
            help="Upload an SBML-qual file for logical models"
        )
    
    with col2:
        st.markdown("### Options")
        
        # Output format selection
        output_csv = st.checkbox(
            "Output as CSV Files",
            value=False,
            key="output_csv",
            help="Output as separate CSV files instead of a single XLSX file"
        )
        
        colon_format = st.checkbox(
            "Use Colon Notation",
            value=False,
            key="colon_format",
            help="Use colon notation (A:2) instead of operators (A >= 2)"
        )
        
        validate_annotations_tab2 = st.checkbox(
            "Validate Annotations",
            value=True,
            key="validate_anno_tab2",
            help="Validate SBML annotations using sbmlutils (requires sbmlutils with metadata.validator)"
        )
        use_name_tab2 = st.checkbox(
            "Use Species Name",
            value=False,
            key="use_name_tab2",
            help="Use Species Name in rules and interactions. If unchecked (default), uses Species_ID."
        )
        
        # # Template options (only for XLSX output)
        # if not output_csv:
        #     use_default_template = st.checkbox(
        #         "Use Default Template",
        #         value=True,
        #         key="use_default_template",
        #         help="Include README and Appendix sheets from default template"
        #     )
            
        #     # Custom template upload
        #     st.markdown("**Or upload custom template:**")
        #     custom_template = st.file_uploader(
        #         "Custom Template (.xlsx)",
        #         type=["xlsx"],
        #         key="custom_template_upload",
        #         help="Upload a custom template file for README and Appendix sheets"
        #     )
        # else:
        #     use_default_template = False
        #     custom_template = None
    
    if uploaded_sbml is not None:
        # Store file content
        file_content = uploaded_sbml.getvalue()
        
        # Preview
        with st.expander("📄 Preview Uploaded SBML", expanded=False):
            try:
                sbml_text = file_content.decode('utf-8')
                if len(sbml_text) > 50000:  # ~50KB
                    preview_lines = sbml_text.split('\n')[:100]
                    st.code('\n'.join(preview_lines) + "\n\n... (truncated)", language="xml")
                else:
                    st.code(sbml_text, language="xml")
                del sbml_text
            except Exception as e:
                st.error(f"Error previewing SBML: {str(e)}")
        
        # Output filename editor
        st.markdown("### Output File Name")
        original_name = uploaded_sbml.name.rsplit('.', 1)[0]
        output_filename = st.text_input(
            "Edit output filename (without extension)",
            value=f"{original_name}_out",
            key="xlsx_output_name",
            help="The file will be saved with .xlsx extension"
        )
        
        # Convert button
        if st.button("🔄 Convert to Spreadsheet", type="primary", key="convert_to_spreadsheet"):
            with st.spinner("Converting..."):
                try:
                    temp_files_to_cleanup = []
                    
                    # Create temporary input file
                    with tempfile.NamedTemporaryFile(suffix='.sbml', delete=False) as tmp_in:
                        tmp_in.write(file_content)
                        input_path = tmp_in.name
                    temp_files_to_cleanup.append(input_path)
                    
                    # Determine template path (only for XLSX) - always use default template
                    template_path = get_default_template_path() if not output_csv else None
                    
                    # Determine rule format
                    rule_format = "colon" if colon_format else "operators"
                    
                    if output_csv:
                        # CSV output - use temp directory
                        temp_dir = tempfile.mkdtemp()
                        output_prefix = os.path.join(temp_dir, output_filename or original_name)
                        
                        # Perform conversion
                        result = convert_sbml_to_spreadsheet(
                            input_path,
                            output_prefix,
                            template_path=None,
                            rule_format=rule_format,
                            output_csv=True,
                            print_messages=False,  # Display in app instead
                            validate=validate_annotations_tab2,
                            use_name=use_name_tab2
                        )
                        
                        created_files = result['created_files']
                        
                        # Create a ZIP file with all CSVs
                        zip_buffer = BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for csv_path in created_files:
                                zip_file.write(csv_path, os.path.basename(csv_path))
                        zip_buffer.seek(0)
                        zip_content = zip_buffer.getvalue()
                        
                        # Success message
                        st.markdown('<div class="success-box">✅ Conversion successful!</div>', unsafe_allow_html=True)
                        
                        # Display validation errors
                        validation_errors = result.get('validation_errors', [])
                        total_val_errors = result.get('total_validation_errors', 0)
                        
                        if validation_errors:
                            with st.expander(f"❗️ Annotation Validation Warnings ({total_val_errors} total)", expanded=True):
                                for error in validation_errors:
                                    st.warning(error)
                        
                        
                        # Display messages
                        all_messages = result.get('warnings', [])
                        if all_messages:
                            info_msgs = [m for m in all_messages if m.startswith("Found ")]
                            use_name_msgs = [m for m in all_messages if "--use-name" in m or "Using ID mode" in m or "Using Name mode" in m]
                            warning_msgs = [m for m in all_messages if m not in info_msgs and m not in use_name_msgs]
                            
                            # Display --use-name related messages in an expander
                            if use_name_msgs:
                                with st.expander(f"ℹ️ --use-name Flag Messages ({len(use_name_msgs)} total)", expanded=False):
                                    for msg in use_name_msgs:
                                        st.info(msg)
                            
                            if warning_msgs:
                                with st.expander(f"⚠️ Warnings ({len(warning_msgs)} total)", expanded=False):
                                    for msg in warning_msgs:
                                        st.info(f"{msg}")
                        
                        # Display statistics
                        species_count = result.get('species', 0)
                        transitions_count = result.get('transitions', 0)
                        interactions_count = result.get('interactions', 0)
                        
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Species", species_count)
                        col2.metric("Transitions", transitions_count)
                        col3.metric("Interactions", interactions_count)
                        
                        st.caption(f"CSV Files: {len(created_files)} | ZIP Size: {len(zip_content)} bytes")
                        
                        # Preview CSV files
                        with st.expander("📊 Preview CSV Output", expanded=False):
                            for csv_path in created_files:
                                st.subheader(f"File: {os.path.basename(csv_path)}")
                                try:
                                    df = pd.read_csv(csv_path, nrows=50)
                                    st.dataframe(df, width='stretch')
                                except Exception as e:
                                    st.error(f"Error previewing: {str(e)}")
                        
                        # Download button - ZIP file
                        final_filename = f"{output_filename or original_name}_csv.zip"
                        st.download_button(
                            label="⬇️ Download CSV Files (ZIP)",
                            data=zip_content,
                            file_name=final_filename,
                            mime="application/zip",
                            type="primary"
                        )
                        
                        # Cleanup
                        import shutil
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    else:
                        # XLSX output
                        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_out:
                            output_path = tmp_out.name
                        temp_files_to_cleanup.append(output_path)
                        
                        # Perform conversion
                        result = convert_sbml_to_spreadsheet(
                            input_path,
                            output_path,
                            template_path=template_path,
                            rule_format=rule_format,
                            output_csv=False,
                            print_messages=False,  # Display in app instead
                            validate=validate_annotations_tab2,
                            use_name=use_name_tab2
                        )

                        created_files = result['created_files']

                        # Read the output file
                        with open(created_files[0], 'rb') as f:
                            xlsx_content = f.read()
                        
                        # Success message
                        st.markdown('<div class="success-box">✅ Conversion successful!</div>', unsafe_allow_html=True)
                        
                        # Display validation errors
                        validation_errors = result.get('validation_errors', [])
                        total_val_errors = result.get('total_validation_errors', 0)
                        if validation_errors:
                            with st.expander(f"❗️ Annotation Validation Warnings ({total_val_errors} total)", expanded=True):
                                for error in validation_errors:
                                    st.warning(error)
                        
                        # Display messages
                        all_messages = result.get('warnings', [])
                        if all_messages:
                            info_msgs = [m for m in all_messages if m.startswith("Found ")]
                            use_name_msgs = [m for m in all_messages if "--use-name" in m or "Using ID mode" in m or "Using Name mode" in m or "enable --use-name" in m or "disable --use-name" in m]
                            warning_msgs = [m for m in all_messages if m not in info_msgs and m not in use_name_msgs]
                            
                            # Display --use-name related messages in an expander
                            if use_name_msgs:
                                with st.expander(f"ℹ️ --use-name Flag Messages ({len(use_name_msgs)} total)", expanded=False):
                                    for msg in use_name_msgs:
                                        st.info(msg)
                            
                            if warning_msgs:
                                with st.expander(f"⚠️ Warnings ({len(warning_msgs)} total)", expanded=False):
                                    for msg in warning_msgs:
                                        st.info(f"{msg}")
                        
                        # Display statistics
                        species_count = result.get('species', 0)
                        transitions_count = result.get('transitions', 0)
                        interactions_count = result.get('interactions', 0)
                        
                        wb = load_workbook(BytesIO(xlsx_content), read_only=True, data_only=True)
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Species", species_count)
                        col2.metric("Transitions", transitions_count)
                        col3.metric("Interactions", interactions_count)
                        
                        st.caption(f"Format: {'Colon' if colon_format else 'Operators'} | File Size: {len(xlsx_content)} bytes")
                        
                        # LIMITED Preview
                        with st.expander("📊 Preview Spreadsheet Output", expanded=False):
                            skip_sheets = ["README", "Appendix"]
                            skipped_sheets = [s for s in wb.sheetnames if s in skip_sheets]
                            
                            if skipped_sheets:
                                st.info(f"ℹ️ Skipping sheets: {', '.join(skipped_sheets)}")
                            
                            for sheet_name in wb.sheetnames:
                                if sheet_name in skip_sheets:
                                    continue
                                
                                st.subheader(f"Sheet: {sheet_name}")
                                sheet = wb[sheet_name]
                                
                                # LIMIT rows for preview
                                data = []
                                max_preview_rows = 50
                                for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                                    if idx >= max_preview_rows:
                                        break
                                    data.append(row)
                                
                                if data:
                                    max_cols = max(len(row) for row in data)
                                    # Convert all values to strings to avoid pyarrow serialization errors
                                    padded_data = [
                                        [str(v) if v is not None else "" for v in list(row) + [None] * (max_cols - len(row))]
                                        for row in data
                                    ]
                                    df = pd.DataFrame(padded_data)
                                    st.dataframe(df, width='stretch')
                                    if idx >= max_preview_rows:
                                        st.info(f"Preview limited to {max_preview_rows} rows")
                                    del df, padded_data, data
                        
                        wb.close()
                        del wb
                        
                        # Download button
                        final_filename = f"{output_filename}.xlsx" if output_filename else f"{original_name}_out.xlsx"
                        st.download_button(
                            label="⬇️ Download Spreadsheet",
                            data=xlsx_content,
                            file_name=final_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                        
                        del xlsx_content
                    
                    # Cleanup temp files
                    for f in temp_files_to_cleanup:
                        try:
                            os.unlink(f)
                        except:
                            pass
                    gc.collect()
                    
                except Exception as e:
                    st.error(f"❌ Conversion failed: {str(e)}")
                    import traceback
                    with st.expander("Error Details"):
                        st.code(traceback.format_exc())
                    gc.collect()

# Footer
st.markdown(f"""
<div style='text-align: center; color: #666; padding: 2rem 0;'>
    <p class="version">TabularQual v{spec.VERSION}</p>
    <p>For issues or feedback, please visit our <a href='https://github.com/sys-bio/TabularQual'>GitHub repository</a></p>
</div>
""", unsafe_allow_html=True)